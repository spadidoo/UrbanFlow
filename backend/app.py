# backend/app.py - COMPLETE VERSION
# Replace your entire app.py with this file
import os
from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from models.predictor import TrafficPredictor
from datetime import datetime, timedelta
import pandas as pd
from services.traffic_api import TrafficAPIService
from services.database import DatabaseService
from services.email_service import send_otp_email
from werkzeug.utils import secure_filename
from flask import send_file
import random
import string
from routes.auth import auth_bp
from flask import jsonify, request, send_file, make_response
from datetime import datetime
import pandas as pd
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from psycopg2.extras import RealDictCursor
from werkzeug.utils import secure_filename
import os
import base64
from datetime import datetime, timedelta
import pytz
import smtplib
from flask import Blueprint, request, jsonify
from flask_mail import Message
import requests



load_dotenv()

app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "http://127.0.0.1:3000"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True
    }
})

# register blueprints after CORS
app.register_blueprint(auth_bp, url_prefix='/api/auth')

# ============================================================
# File upload configuration
# Path to preprocessed data folder - adjust this to match your structure
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'data', 'processed')
ALLOWED_EXTENSIONS = {'csv'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max




# Ensure folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_csv_info(filepath):
    """Get CSV file information (rows, columns)"""
    try:
        df = pd.read_csv(filepath)
        return {
            'rows': len(df),
            'cols': len(df.columns)
        }
    except Exception as e:
        print(f"Error reading CSV {filepath}: {e}")
        return {
            'rows': 0,
            'cols': 0
        }
    


AVATAR_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'uploads', 'avatars')
AVATAR_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE = 2 * 1024 * 1024

# ============================================================

print("Loading trained model...")
predictor = TrafficPredictor()
print("✓ Model loaded and ready!")

# After creating the app
traffic_service = TrafficAPIService()

print("Loading traffic API service...")
traffic_service = TrafficAPIService()
print("✓ Traffic API service ready!")

print("Loading database service...")
db = DatabaseService()
print("✓ Database service ready!")


os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(AVATAR_FOLDER, exist_ok=True)

def allowed_csv_file(filename):
    """Check if file has allowed CSV extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in {'csv'}

def allowed_avatar_file(filename):
    """Check if file has allowed image extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in AVATAR_EXTENSIONS

def save_avatar_to_db(user_id, image_data, filename):
    """Save avatar image data to database"""
    conn = None
    try:
        conn = db._get_connection()
        cursor = conn.cursor()
        
        # Update user's avatar in database
        # Store as base64 in profile_photo column
        cursor.execute("""
            UPDATE users 
            SET profile_photo = %s,
                profile_photo_filename = %s,
                updated_at = NOW()
            WHERE user_id = %s
        """, (image_data, filename, user_id))
        
        conn.commit()
        print(f"✓ Avatar saved for user {user_id}")
        return True
        
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"✗ Error saving avatar: {e}")
        return False
        
    finally:
        if conn:
            conn.close()

contact_bp = Blueprint('contact', __name__)

@contact_bp.route('/api/contact', methods=['POST'])
def contact_form():
    try:
        data = request.get_json()
        
        # Validate required fields
        errors = {}
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        subject = data.get('subject', '').strip()
        message = data.get('message', '').strip()
        
        if not name:
            errors['name'] = 'Name is required'
        if not email:
            errors['email'] = 'Email is required'
        elif '@' not in email or '.' not in email:
            errors['email'] = 'Please enter a valid email address'
        if not subject:
            errors['subject'] = 'Subject is required'
        if not message:
            errors['message'] = 'Message is required'
        
        if errors:
            return jsonify({'success': False, 'errors': errors}), 400
        
        # Send email using smtplib (simpler, no Flask-Mail needed)
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        smtp_email = os.getenv('SMTP_EMAIL')
        smtp_password = os.getenv('SMTP_PASSWORD')
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'[UrbanFlow Contact] {subject}'
        msg['From'] = smtp_email
        msg['To'] = smtp_email  # Send to yourself
        msg['Reply-To'] = email
        
        # Plain text version
        text_content = f"""
New contact form submission from UrbanFlow:

Name: {name}
Email: {email}
Subject: {subject}

Message:
{message}
        """
        
        # HTML version
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #2563eb;">New Contact Form Submission</h2>
            <p style="color: #666;">You have received a new message from the UrbanFlow website.</p>
            
            <table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold; width: 100px;">Name:</td>
                    <td style="padding: 10px; border-bottom: 1px solid #eee;">{name}</td>
                </tr>
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold;">Email:</td>
                    <td style="padding: 10px; border-bottom: 1px solid #eee;">
                        <a href="mailto:{email}">{email}</a>
                    </td>
                </tr>
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #eee; font-weight: bold;">Subject:</td>
                    <td style="padding: 10px; border-bottom: 1px solid #eee;">{subject}</td>
                </tr>
            </table>
            
            <div style="margin-top: 20px; padding: 15px; background-color: #f9fafb; border-radius: 8px;">
                <p style="font-weight: bold; margin-bottom: 10px;">Message:</p>
                <p style="white-space: pre-wrap; margin: 0;">{message}</p>
            </div>
            
            <hr style="margin-top: 30px; border: none; border-top: 1px solid #eee;" />
            <p style="color: #999; font-size: 12px;">
                This email was sent from the UrbanFlow contact form.
            </p>
        </div>
        """
        
        msg.attach(MIMEText(text_content, 'plain'))
        msg.attach(MIMEText(html_content, 'html'))
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.send_message(msg)
        
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Contact form error: {e}")
        return jsonify({
            'success': False, 
            'error': 'Failed to send message. Please try again later.'
        }), 500


@app.route('/api/google-directions', methods=['POST'])
def google_directions():
    """Proxy endpoint for Google Directions API to avoid CORS"""
    try:
        data = request.get_json()
        origin = data.get('origin')  # "lat,lng"
        destination = data.get('destination')  # "lat,lng"
        
        if not origin or not destination:
            return jsonify({'error': 'Missing origin or destination'}), 400
        
        # Get API key from environment or hardcode temporarily
        api_key = os.environ.get('GOOGLE_MAPS_API_KEY', 'YOUR_API_KEY_HERE')
        
        url = f"https://maps.googleapis.com/maps/api/directions/json?origin={origin}&destination={destination}&mode=driving&key={api_key}"
        
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            return jsonify(response.json())
        else:
            return jsonify({'error': f'Google API error: {response.status_code}'}), response.status_code
            
    except Exception as e:
        print(f"❌ Google Directions error: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================
# NEW ROUTE: Upload Avatar
# ============================================================

@app.route('/api/upload-avatar', methods=['POST'])
def upload_avatar():
    """
    Upload user avatar/profile picture
    
    Form data:
        - avatar: Image file
        - user_id: User ID
    
    Returns:
        - success: Boolean
        - avatar_url: URL to access the avatar (if successful)
        - error: Error message (if failed)
    """
    try:
        # Check if file is in request
        if 'avatar' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['avatar']
        
        # Check if file was selected
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400
        
        # Get user_id
        user_id = request.form.get('user_id')
        if not user_id:
            return jsonify({
                'success': False,
                'error': 'user_id is required'
            }), 400
        
        user_id = int(user_id)
        
        # Validate file type
        if not allowed_avatar_file(file.filename):
            return jsonify({
                'success': False,
                'error': 'Invalid file type. Allowed: PNG, JPG, JPEG, GIF, WEBP'
            }), 400
        
        # Read file data
        file_data = file.read()
        
        # Check file size
        if len(file_data) > MAX_FILE_SIZE:
            return jsonify({
                'success': False,
                'error': f'File too large. Maximum size: {MAX_FILE_SIZE / (1024*1024)}MB'
            }), 400
        
        # Generate secure filename
        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"user_{user_id}_{timestamp}_{original_filename}"
        
        # Convert to base64 for database storage
        base64_data = base64.b64encode(file_data).decode('utf-8')
        
        # Get file extension for data URL
        file_ext = original_filename.rsplit('.', 1)[1].lower()
        mime_type = f"image/{file_ext if file_ext != 'jpg' else 'jpeg'}"
        
        # Create data URL for storage
        data_url = f"data:{mime_type};base64,{base64_data}"
        
        # Save to database
        success = save_avatar_to_db(user_id, data_url, filename)
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Failed to save avatar to database'
            }), 500
        
        # OPTIONAL: Also save to filesystem
        # file_path = os.path.join(UPLOAD_FOLDER, filename)
        # with open(file_path, 'wb') as f:
        #     f.write(file_data)
        
        return jsonify({
            'success': True,
            'avatar_url': data_url,  # Return data URL for immediate preview
            'filename': filename,
            'message': 'Avatar uploaded successfully'
        })
        
    except Exception as e:
        import traceback
        print(f"✗ Error uploading avatar: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# NEW ROUTE: Get User Avatar
# ============================================================

@app.route('/api/user/<int:user_id>/avatar', methods=['GET'])
def get_user_avatar(user_id):
    """
    Get user's avatar image
    
    Returns the avatar as base64 data URL
    """
    conn = None
    try:
        conn = db._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT profile_photo, profile_photo_filename
            FROM users
            WHERE user_id = %s
        """, (user_id,))
        
        result = cursor.fetchone()
        
        if not result or not result[0]:
            return jsonify({
                'success': False,
                'error': 'No avatar found'
            }), 404
        
        return jsonify({
            'success': True,
            'avatar_url': result[0],
            'filename': result[1]
        })
        
    except Exception as e:
        print(f"✗ Error fetching avatar: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
    finally:
        if conn:
            conn.close()

@app.route('/api/user/profile', methods=['GET'])
def get_user_profile():
    """
    Get current user's profile including avatar
    """
    try:
        # Get user_id from auth (for now, using hardcoded)
        user_id = request.args.get('user_id', type=int)  # Your user is ID 3
        
        conn = db._get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                user_id,
                username,
                email,
                first_name,
                last_name,
                role,
                profile_photo,          -- CRITICAL: Include this
                profile_photo_filename,
                organization,
                created_at,
                updated_at
            FROM users
            WHERE user_id = %s
        """, (user_id,))
        
        user = cursor.fetchone()
        
        if not user:
            return jsonify({
                'success': False,
                'error': 'User not found'
            }), 404
        
        # Convert to dict and format
        user_data = dict(user)
        
        return jsonify({
            'success': True,
            'user': {
                'id': user_data['user_id'],
                'username': user_data['username'],
                'email': user_data['email'],
                'firstName': user_data['first_name'],
                'lastName': user_data['last_name'],
                'role': user_data['role'],
                'name': f"{user_data['first_name']} {user_data['last_name']}",
                'avatarUrl': user_data['profile_photo'] or '/urban_planner_icon.png',  # CRITICAL
                'organization': user_data.get('organization')
            }
        })
        
    except Exception as e:
        print(f"Error fetching profile: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if conn:
            conn.close()

# ============================================================
# NEW ROUTE: Get Completed/Finished Simulations (for Reports)
# ============================================================

@app.route('/api/reports', methods=['GET'])
def get_finished_reports():
    """
    Get all finished simulations for the reports page
    
    A simulation is considered finished when:
    - Its end_time is before current time
    - OR simulation_status is 'completed' or 'published'
    
    Query params:
        - query: Search by title or location
        - date: Filter by date (YYYY-MM-DD)
        - location: Filter by location
        - type: Filter by disruption type
        - user_id: User ID (optional, defaults to 2)
    """
    try:
        # Get query parameters
        search_query = request.args.get('query', '').strip()
        filter_date = request.args.get('date', '').strip()
        filter_location = request.args.get('location', '').strip()
        filter_type = request.args.get('type', '').strip()
        user_id = request.args.get('user_id', type=int)
        
        conn = None
        try:
            conn = db._get_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Base query for finished simulations
            query = """
                SELECT 
                    sr.simulation_id,
                    sr.simulation_name,
                    sr.description,
                    sr.disruption_type,
                    sr.disruption_location,
                    sr.start_time,
                    sr.end_time,
                    sr.severity_level,
                    sr.simulation_status,
                    sr.total_affected_segments,
                    sr.average_delay_ratio,
                    sr.created_at,
                    sr.updated_at,
                    pr.published_id,
                    pr.title as published_title,
                    pr.published_at,
                    pr.is_active as is_published
                FROM simulation_runs sr
                LEFT JOIN published_runs pr ON sr.simulation_id = pr.simulation_id
                WHERE sr.user_id = %s
                AND sr.simulation_status != 'deleted'
                AND (
                    sr.end_time < NOW()
                    OR sr.simulation_status IN ('completed', 'published')
                )
            """
            
            params = [user_id]
            
            # Apply filters
            if search_query:
                query += """ AND (
                    LOWER(sr.simulation_name) LIKE %s 
                    OR LOWER(sr.disruption_location) LIKE %s
                    OR LOWER(pr.title) LIKE %s
                )"""
                search_pattern = f'%{search_query.lower()}%'
                params.extend([search_pattern, search_pattern, search_pattern])
            
            if filter_date:
                query += " AND DATE(sr.start_time) = %s"
                params.append(filter_date)
            
            if filter_location:
                query += " AND LOWER(sr.disruption_location) LIKE %s"
                params.append(f'%{filter_location.lower()}%')
            
            if filter_type and filter_type.lower() != 'all':
                query += " AND LOWER(sr.disruption_type) = %s"
                params.append(filter_type.lower())
            
            query += " ORDER BY sr.end_time DESC, sr.created_at DESC"
            
            cursor.execute(query, params)
            simulations = cursor.fetchall()
            
            # Transform to frontend format
            reports = []
            for sim in simulations:
                # Determine the display title
                title = sim.get('published_title') or sim['simulation_name']
                
                # Format dates
                start_date = sim['start_time'].strftime('%Y-%m-%d') if sim['start_time'] else 'N/A'
                end_date = sim['end_time'].strftime('%Y-%m-%d') if sim['end_time'] else 'N/A'
                date_range = f"{start_date} to {end_date}" if start_date != end_date else start_date
                
                # Extract barangay from location
                location = sim['disruption_location'] or 'Unknown Location'
                
                reports.append({
                    'id': sim['simulation_id'],
                    'title': title,
                    'location': location,
                    'date': date_range,
                    'start_date': start_date,
                    'end_date': end_date,
                    'type': (sim['disruption_type'] or 'general').capitalize(),
                    'severity_level': sim['severity_level'] or 'moderate',
                    'status': 'Published' if sim.get('is_published') else 'Completed',
                    'total_affected_segments': sim['total_affected_segments'] or 0,
                    'average_delay_ratio': float(sim['average_delay_ratio']) if sim['average_delay_ratio'] else 0.0,
                    'created_at': sim['created_at'].isoformat() if sim['created_at'] else None,
                })
            
            return jsonify({
                'success': True,
                'reports': reports,
                'count': len(reports)
            })
            
        except Exception as e:
            print(f"✗ Error retrieving reports: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
            
        finally:
            if conn:
                conn.close()
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# NEW ROUTE: Export Report as PDF
# ============================================================

@app.route('/api/reports/<int:simulation_id>/export', methods=['GET'])
def export_report(simulation_id):
    """
    Export a report in PDF, CSV, or Excel format
    
    Query params:
        - format: 'pdf', 'csv', or 'excel' (required)
    """
    try:
        export_format = request.args.get('format', '').lower()
        
        if export_format not in ['pdf', 'csv', 'excel']:
            return jsonify({
                'success': False,
                'error': 'Invalid format. Must be pdf, csv, or excel'
            }), 400
        
        # Get simulation data
        simulation = db.get_simulation_by_id(simulation_id)
        
        if not simulation:
            return jsonify({
                'success': False,
                'error': 'Simulation not found'
            }), 404
        
        # Generate export based on format
        if export_format == 'pdf':
            return export_as_pdf(simulation)
        elif export_format == 'csv':
            return export_as_csv(simulation)
        elif export_format == 'excel':
            return export_as_excel(simulation)
            
    except Exception as e:
        import traceback
        print(f"✗ Error exporting report: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def export_as_pdf(simulation):
    """
    Generate COMPACT PDF export with minimal spacing
    All sections flow continuously without unnecessary page breaks
    """
    buffer = io.BytesIO()
    
    # ============================================================
    # COMPACT DOCUMENT SETUP - Reduced margins
    # ============================================================
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=40,      # Reduced from 50
        leftMargin=40,       # Reduced from 50
        topMargin=35,        # Reduced from 50
        bottomMargin=25      # Reduced from 30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ============================================================
    # COMPACT STYLES - Reduced spacing
    # ============================================================
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,          # Reduced from 22
        textColor=colors.HexColor('#FF6B35'),
        spaceAfter=8,         # Reduced from 20
        spaceBefore=0,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,          # Reduced from 14
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=6,         # Reduced from 10
        spaceBefore=8,        # Reduced from 15
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubheading',
        parent=styles['Heading3'],
        fontSize=10,          # Reduced from 12
        textColor=colors.HexColor('#34495E'),
        spaceAfter=4,         # Reduced from 8
        spaceBefore=6,        # Reduced from 10
        fontName='Helvetica-Bold'
    )
    
    compact_normal = ParagraphStyle(
        'CompactNormal',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        spaceBefore=0
    )
    
    # ============================================================
    # TITLE AND METADATA - Compact
    # ============================================================
    
    title = Paragraph(
        f"UrbanFlow Traffic Simulation Report<br/>{simulation['simulation_name']}", 
        title_style
    )
    elements.append(title)
    elements.append(Spacer(1, 6))  # Reduced from 10
    
    # Report metadata - compact
    metadata_text = f"""
    <b>Report Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>Simulation ID:</b> {simulation['simulation_id']}<br/>
    <b>Report Type:</b> Traffic Disruption Analysis
    """
    elements.append(Paragraph(metadata_text, compact_normal))
    elements.append(Spacer(1, 8))  # Reduced from 15
    
    # ============================================================
    # BASIC INFORMATION TABLE - Compact
    # ============================================================
    
    elements.append(Paragraph("Simulation Information", heading_style))
    
    info_data = [
        ['Field', 'Value'],
        ['Type', (simulation['disruption_type'] or 'N/A').capitalize()],
        ['Location', simulation['disruption_location'] or 'N/A'],
        ['Start Date', simulation['start_time'].strftime('%Y-%m-%d %H:%M') if simulation['start_time'] else 'N/A'],
        ['End Date', simulation['end_time'].strftime('%Y-%m-%d %H:%M') if simulation['end_time'] else 'N/A'],
        ['Severity Level', (simulation['severity_level'] or 'N/A').capitalize()],
        ['Duration', f"{calculate_duration_hours(simulation)} hours"],
    ]
    
    info_table = Table(info_data, colWidths=[2.2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Reduced from 10
        ('TOPPADDING', (0, 0), (-1, 0), 6),     # Reduced from 10
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#ECF0F1')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('TOPPADDING', (0, 1), (-1, -1), 4),    # Reduced padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 10))  # Reduced from 20
    
    # ============================================================
    # SUMMARY METRICS - Compact
    # ============================================================
    
    elements.append(Paragraph("Key Performance Indicators", heading_style))
    
    summary_data = [
        ['Metric', 'Value', 'Impact Level'],
        ['Total Affected Segments', str(simulation.get('total_affected_segments', 'N/A')), '-'],
        ['Average Delay Ratio', f"{simulation.get('average_delay_ratio', 0):.2f}x", get_impact_level(simulation.get('average_delay_ratio', 0))],
        ['Peak Congestion Hour', get_peak_hour(simulation), '-'],
        ['Status', simulation.get('simulation_status', 'N/A').capitalize(), '-'],
    ]
    
    summary_table = Table(summary_data, colWidths=[2.2*inch, 1.8*inch, 1.8*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 10))  # Reduced from 20
    
    # ============================================================
    # HOUR-BY-HOUR BREAKDOWN - NO PAGE BREAK, COMPACT
    # ============================================================
    
    if simulation.get('results') and len(simulation['results']) > 0:
        # NO PageBreak() here - let it flow naturally
        elements.append(Paragraph("Hour-by-Hour Analysis", heading_style))
        elements.append(Paragraph(
            "Detailed breakdown of traffic conditions for each hour.",
            compact_normal
        ))
        elements.append(Spacer(1, 6))  # Reduced from 10
        
        # Prepare hourly data - limit to reasonable amount per page
        hourly_data = [['Hour', 'Time', 'Severity', 'Delay (min)', 'Status']]
        
        # Show up to 24 hours on first page, then continue if needed
        max_hours_first = 24
        for i, result in enumerate(simulation['results'][:max_hours_first]):
            hour = result.get('hour', i)
            severity = result.get('severity', 0)
            delay_minutes = result.get('delay_minutes', 0)
            
            if simulation['start_time']:
                time_obj = simulation['start_time'] + timedelta(hours=hour)
                time_str = time_obj.strftime('%I:%M %p')
            else:
                time_str = f"Hour {hour}"
            
            status = get_severity_status(severity)
            
            hourly_data.append([
                str(hour),
                time_str,
                f"{severity:.2f}",
                str(delay_minutes),
                status
            ])
        
        # Compact hourly table
        hourly_table = Table(hourly_data, colWidths=[0.5*inch, 1.2*inch, 0.9*inch, 1*inch, 1.2*inch])
        
        # Compact table style
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),  # Reduced
            ('TOPPADDING', (0, 0), (-1, 0), 5),     # Reduced
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 3),    # Reduced
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3), # Reduced
        ]
        
        # Color rows based on severity
        for i, result in enumerate(simulation['results'][:max_hours_first], start=1):
            severity = result.get('severity', 0)
            if severity >= 2.0:
                bg_color = colors.HexColor('#FADBD8')
            elif severity >= 1.5:
                bg_color = colors.HexColor('#FCF3CF')
            elif severity >= 1.0:
                bg_color = colors.HexColor('#D5F4E6')
            else:
                bg_color = colors.white
            
            table_style.append(('BACKGROUND', (0, i), (-1, i), bg_color))
        
        hourly_table.setStyle(TableStyle(table_style))
        elements.append(hourly_table)
        elements.append(Spacer(1, 8))  # Reduced from 15
        
        # Legend - compact
        legend_text = """
        <b>Severity Legend:</b> Light (&lt;1.5), Moderate (1.5-2.0), Heavy (&gt;2.0)
        """
        elements.append(Paragraph(legend_text, compact_normal))
        
        # ============================================================
        # DAY-BY-DAY BREAKDOWN - NO PAGE BREAK if multi-day
        # ============================================================
        
        duration_hours = calculate_duration_hours(simulation)
        if duration_hours > 24:
            # NO PageBreak() - let it continue on same page if space allows
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("Day-by-Day Summary", heading_style))
            elements.append(Paragraph(
                "Daily aggregated metrics showing average conditions.",
                compact_normal
            ))
            elements.append(Spacer(1, 6))
            
            daily_summary = calculate_daily_summary(simulation)
            
            daily_data = [['Day', 'Date', 'Avg Severity', 'Peak Hour', 'Max Delay (min)', 'Status']]
            
            for day_info in daily_summary[:14]:  # Limit to 14 days per page
                daily_data.append([
                    f"Day {day_info['day_number']}",
                    day_info['date'],
                    f"{day_info['avg_severity']:.2f}",
                    day_info['peak_hour'],
                    str(day_info['max_delay']),
                    day_info['status']
                ])
            
            daily_table = Table(daily_data, colWidths=[0.6*inch, 1.3*inch, 1*inch, 1*inch, 1.1*inch, 1*inch])
            daily_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ]))
            
            elements.append(daily_table)
            elements.append(Spacer(1, 8))
        
        # ============================================================
        # WEEK-BY-WEEK BREAKDOWN - Only if very long
        # ============================================================
        
        if duration_hours > 168:  # More than 1 week
            # Add page break only for weekly summary since it's a separate analysis
            elements.append(PageBreak())
            elements.append(Paragraph("Week-by-Week Summary", heading_style))
            elements.append(Paragraph(
                "Weekly aggregated trends showing overall patterns.",
                compact_normal
            ))
            elements.append(Spacer(1, 6))
            
            weekly_summary = calculate_weekly_summary(simulation)
            
            weekly_data = [['Week', 'Date Range', 'Avg Severity', 'Total Hours', 'Avg Delay (min)', 'Trend']]
            
            for week_info in weekly_summary:
                weekly_data.append([
                    f"Week {week_info['week_number']}",
                    week_info['date_range'],
                    f"{week_info['avg_severity']:.2f}",
                    str(week_info['total_hours']),
                    f"{week_info['avg_delay']:.1f}",
                    week_info['trend']
                ])
            
            weekly_table = Table(weekly_data, colWidths=[0.7*inch, 1.8*inch, 1*inch, 1*inch, 1.1*inch, 0.9*inch])
            weekly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16A085')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                ('TOPPADDING', (0, 1), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ]))
            
            elements.append(weekly_table)
            elements.append(Spacer(1, 8))
    
    # ============================================================
    # FOOTER - Compact
    # ============================================================
    
    elements.append(Spacer(1, 15))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1
    )
    footer_text = f"""
    <b>UrbanFlow Traffic Simulation System</b><br/>
    Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
    For official use by CCPOSO and DPWH | Calamba City, Laguna
    """
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    
    # Prepare response
    buffer.seek(0)
    filename = f"urbanflow_report_{simulation['simulation_id']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    from flask import make_response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_as_csv(simulation):
    """Generate CSV export of simulation report"""
    # Prepare data
    data = {
        'Simulation ID': [simulation['simulation_id']],
        'Title': [simulation['simulation_name']],
        'Type': [(simulation['disruption_type'] or 'N/A').capitalize()],
        'Location': [simulation['disruption_location'] or 'N/A'],
        'Start Date': [simulation['start_time'].strftime('%Y-%m-%d %H:%M') if simulation['start_time'] else 'N/A'],
        'End Date': [simulation['end_time'].strftime('%Y-%m-%d %H:%M') if simulation['end_time'] else 'N/A'],
        'Severity Level': [(simulation['severity_level'] or 'N/A').capitalize()],
        'Total Affected Segments': [simulation.get('total_affected_segments', 'N/A')],
        'Average Delay Ratio': [simulation.get('average_delay_ratio', 'N/A')],
        'Status': [simulation.get('simulation_status', 'N/A').capitalize()],
    }
    
    df = pd.DataFrame(data)
    
    # If there are results, add them as well
    if simulation.get('results') and len(simulation['results']) > 0:
        # Create a separate section for hourly results
        results_data = []
        for result in simulation['results']:
            results_data.append({
                'Hour': result.get('hour', 'N/A'),
                'Severity': result.get('severity', 'N/A'),
                'Delay Minutes': result.get('delay_minutes', 'N/A')
            })
        
        results_df = pd.DataFrame(results_data)
        
        # Combine both dataframes with a separator
        output = io.StringIO()
        df.to_csv(output, index=False)
        output.write('\n\n--- Hourly Results ---\n')
        results_df.to_csv(output, index=False)
        csv_content = output.getvalue()
    else:
        csv_content = df.to_csv(index=False)
    
    # Prepare response
    filename = f"report_{simulation['simulation_id']}_{simulation['simulation_name'].replace(' ', '_')}.csv"
    
    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_as_excel(simulation):
    """Generate Excel export of simulation report"""
    output = io.BytesIO()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add header
    ws_summary['A1'] = 'UrbanFlow Traffic Simulation Report'
    ws_summary['A1'].font = Font(size=16, bold=True, color='FF6B35')
    ws_summary.merge_cells('A1:B1')
    
    # Add basic info
    info_data = [
        ['Simulation ID', simulation['simulation_id']],
        ['Title', simulation['simulation_name']],
        ['Type', (simulation['disruption_type'] or 'N/A').capitalize()],
        ['Location', simulation['disruption_location'] or 'N/A'],
        ['Start Date', simulation['start_time'].strftime('%Y-%m-%d %H:%M') if simulation['start_time'] else 'N/A'],
        ['End Date', simulation['end_time'].strftime('%Y-%m-%d %H:%M') if simulation['end_time'] else 'N/A'],
        ['Severity Level', (simulation['severity_level'] or 'N/A').capitalize()],
        ['Total Affected Segments', simulation.get('total_affected_segments', 'N/A')],
        ['Average Delay Ratio', simulation.get('average_delay_ratio', 'N/A')],
        ['Status', simulation.get('simulation_status', 'N/A').capitalize()],
    ]
    
    row = 3
    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Style the summary sheet
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    
    # Sheet 2: Hourly Results (if available)
    if simulation.get('results') and len(simulation['results']) > 0:
        ws_results = wb.create_sheet(title="Hourly Results")
        
        # Add headers
        headers = ['Hour', 'Severity', 'Delay (minutes)']
        for col, header in enumerate(headers, start=1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row, result in enumerate(simulation['results'], start=2):
            ws_results.cell(row=row, column=1, value=result.get('hour', 'N/A'))
            ws_results.cell(row=row, column=2, value=result.get('severity', 'N/A'))
            ws_results.cell(row=row, column=3, value=result.get('delay_minutes', 'N/A'))
        
        # Style columns
        ws_results.column_dimensions['A'].width = 15
        ws_results.column_dimensions['B'].width = 15
        ws_results.column_dimensions['C'].width = 20
    
    # Save workbook
    wb.save(output)
    output.seek(0)
    
    # Prepare response
    filename = f"report_{simulation['simulation_id']}_{simulation['simulation_name'].replace(' ', '_')}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def calculate_duration_hours(simulation):
    """Calculate total duration in hours"""
    if simulation.get('start_time') and simulation.get('end_time'):
        delta = simulation['end_time'] - simulation['start_time']
        return int(delta.total_seconds() / 3600)
    return len(simulation.get('results', []))

def get_peak_hour(simulation):
    """Find the hour with highest severity"""
    results = simulation.get('results', [])
    if not results:
        return 'N/A'
    
    max_result = max(results, key=lambda x: x.get('severity', 0))
    hour = max_result.get('hour', 0)
    
    if simulation.get('start_time'):
        time_obj = simulation['start_time'] + timedelta(hours=hour)
        return time_obj.strftime('%I:%M %p')
    
    return f"Hour {hour}"

def get_impact_level(delay_ratio):
    """Get impact level from delay ratio"""
    if delay_ratio >= 2.0:
        return 'High'
    elif delay_ratio >= 1.5:
        return 'Moderate'
    elif delay_ratio >= 1.0:
        return 'Low'
    else:
        return 'Minimal'

def get_severity_status(severity):
    """Get status label from severity"""
    if severity >= 2.0:
        return 'Heavy'
    elif severity >= 1.5:
        return 'Moderate'
    elif severity >= 1.0:
        return 'Light'
    else:
        return 'Normal'

def calculate_daily_summary(simulation):
    """Calculate day-by-day summary with accurate peak hour detection"""
    results = simulation.get('results', [])
    start_time = simulation.get('start_time')
    
    if not results or not start_time:
        return []
    
    daily_summary = []
    current_day = []
    day_number = 1
    
    for i, result in enumerate(results):
        hour = result.get('hour', i)
        
        # Group by 24-hour periods
        if hour > 0 and hour % 24 == 0:
            if current_day:
                # ✅ FIXED: Accurate peak hour detection
                # Find all hours with severity within 10% of max severity
                max_severity = max(r.get('severity', 0) for r in current_day)
                threshold = max_severity * 0.9  # Within 90% of max
                
                peak_hours = [
                    r.get('hour', 0) % 24 
                    for r in current_day 
                    if r.get('severity', 0) >= threshold
                ]
                
                # Format peak hours
                if len(peak_hours) == 1:
                    peak_hour_str = f"{peak_hours[0]:02d}:00"
                elif len(peak_hours) <= 3:
                    peak_hour_str = ", ".join(f"{h:02d}:00" for h in sorted(peak_hours))
                else:
                    # Show range if many consecutive hours
                    peak_hours.sort()
                    peak_hour_str = f"{peak_hours[0]:02d}:00 - {peak_hours[-1]:02d}:00"
                
                # ✅ FIXED: Calculate average delay during peak hours only
                peak_delays = [
                    r.get('delay_minutes', 0) 
                    for r in current_day 
                    if (r.get('hour', 0) % 24) in peak_hours
                ]
                avg_peak_delay = sum(peak_delays) / len(peak_delays) if peak_delays else 0
                
                # Calculate overall averages
                avg_severity = sum(r.get('severity', 0) for r in current_day) / len(current_day)
                max_delay = max(r.get('delay_minutes', 0) for r in current_day)
                
                date_obj = start_time + timedelta(days=day_number-1)
                
                daily_summary.append({
                    'day_number': day_number,
                    'date': date_obj.strftime('%Y-%m-%d'),
                    'avg_severity': round(avg_severity, 2),
                    'peak_hour': peak_hour_str,  # ✅ Now shows range
                    'peak_hours': peak_hours,  # ✅ List of peak hours
                    'max_delay': max_delay,
                    'avg_peak_delay': round(avg_peak_delay),  # ✅ NEW: Avg delay in peak hours
                    'status': get_severity_status(avg_severity)
                })
                
                current_day = []
                day_number += 1
        
        current_day.append(result)
    
    # Handle last day
    if current_day:
        max_severity = max(r.get('severity', 0) for r in current_day)
        threshold = max_severity * 0.9
        
        peak_hours = [
            r.get('hour', 0) % 24 
            for r in current_day 
            if r.get('severity', 0) >= threshold
        ]
        
        if len(peak_hours) == 1:
            peak_hour_str = f"{peak_hours[0]:02d}:00"
        elif len(peak_hours) <= 3:
            peak_hour_str = ", ".join(f"{h:02d}:00" for h in sorted(peak_hours))
        else:
            peak_hours.sort()
            peak_hour_str = f"{peak_hours[0]:02d}:00 - {peak_hours[-1]:02d}:00"
        
        peak_delays = [
            r.get('delay_minutes', 0) 
            for r in current_day 
            if (r.get('hour', 0) % 24) in peak_hours
        ]
        avg_peak_delay = sum(peak_delays) / len(peak_delays) if peak_delays else 0
        
        avg_severity = sum(r.get('severity', 0) for r in current_day) / len(current_day)
        max_delay = max(r.get('delay_minutes', 0) for r in current_day)
        
        date_obj = start_time + timedelta(days=day_number-1)
        
        daily_summary.append({
            'day_number': day_number,
            'date': date_obj.strftime('%Y-%m-%d'),
            'avg_severity': round(avg_severity, 2),
            'peak_hour': peak_hour_str,
            'peak_hours': peak_hours,
            'max_delay': max_delay,
            'avg_peak_delay': round(avg_peak_delay),
            'status': get_severity_status(avg_severity)
        })
    
    return daily_summary

def calculate_weekly_summary(simulation):
    """Calculate week-by-week summary with accurate peak detection"""
    results = simulation.get('results', [])
    start_time = simulation.get('start_time')
    
    if not results or not start_time:
        return []
    
    weekly_summary = []
    current_week = []
    week_number = 1
    
    for i, result in enumerate(results):
        hour = result.get('hour', i)
        
        # Group by 168-hour periods (7 days)
        if hour > 0 and hour % 168 == 0:
            if current_week:
                # ✅ FIXED: Calculate accurate weekly metrics
                avg_severity = sum(r.get('severity', 0) for r in current_week) / len(current_week)
                avg_delay = sum(r.get('delay_minutes', 0) for r in current_week) / len(current_week)
                
                # Find peak hours
                max_severity = max(r.get('severity', 0) for r in current_week)
                threshold = max_severity * 0.9
                
                peak_hours = [
                    r for r in current_week 
                    if r.get('severity', 0) >= threshold
                ]
                
                # Calculate average delay during peak hours
                avg_peak_delay = sum(r.get('delay_minutes', 0) for r in peak_hours) / len(peak_hours) if peak_hours else 0
                
                week_start = start_time + timedelta(days=(week_number-1)*7)
                week_end = week_start + timedelta(days=6)
                
                # Determine trend
                first_half_avg = sum(r.get('severity', 0) for r in current_week[:len(current_week)//2]) / (len(current_week)//2)
                second_half_avg = sum(r.get('severity', 0) for r in current_week[len(current_week)//2:]) / (len(current_week) - len(current_week)//2)
                
                if second_half_avg > first_half_avg + 0.2:
                    trend = '↑ Rising'
                elif second_half_avg < first_half_avg - 0.2:
                    trend = '↓ Falling'
                else:
                    trend = '→ Stable'
                
                weekly_summary.append({
                    'week_number': week_number,
                    'date_range': f"{week_start.strftime('%m/%d')} - {week_end.strftime('%m/%d')}",
                    'start_date': week_start.strftime('%Y-%m-%d'),
                    'end_date': week_end.strftime('%Y-%m-%d'),
                    'avg_severity': round(avg_severity, 2),
                    'total_hours': len(current_week),
                    'avg_delay': round(avg_delay, 1),
                    'avg_peak_delay': round(avg_peak_delay, 1),  # ✅ NEW
                    'trend': trend
                })
                
                current_week = []
                week_number += 1
        
        current_week.append(result)
    
    # Handle last week (same fix as above)
    if current_week:
        avg_severity = sum(r.get('severity', 0) for r in current_week) / len(current_week)
        avg_delay = sum(r.get('delay_minutes', 0) for r in current_week) / len(current_week)
        
        max_severity = max(r.get('severity', 0) for r in current_week)
        threshold = max_severity * 0.9
        peak_hours = [r for r in current_week if r.get('severity', 0) >= threshold]
        avg_peak_delay = sum(r.get('delay_minutes', 0) for r in peak_hours) / len(peak_hours) if peak_hours else 0
        
        week_start = start_time + timedelta(days=(week_number-1)*7)
        week_end = start_time + timedelta(hours=len(results))
        
        first_half_avg = sum(r.get('severity', 0) for r in current_week[:len(current_week)//2]) / max(len(current_week)//2, 1)
        second_half_avg = sum(r.get('severity', 0) for r in current_week[len(current_week)//2:]) / max(len(current_week) - len(current_week)//2, 1)
        
        if second_half_avg > first_half_avg + 0.2:
            trend = '↑ Rising'
        elif second_half_avg < first_half_avg - 0.2:
            trend = '↓ Falling'
        else:
            trend = '→ Stable'
        
        weekly_summary.append({
            'week_number': week_number,
            'date_range': f"{week_start.strftime('%m/%d')} - {week_end.strftime('%m/%d')}",
            'start_date': week_start.strftime('%Y-%m-%d'),
            'end_date': week_end.strftime('%Y-%m-%d'),
            'avg_severity': round(avg_severity, 2),
            'total_hours': len(current_week),
            'avg_delay': round(avg_delay, 1),
            'avg_peak_delay': round(avg_peak_delay, 1),  # ✅ NEW
            'trend': trend
        })
    
    return weekly_summary

# health check endpoint
@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint for dashboard"""
    try:
        # Test database connection
        conn = db._get_connection()
        conn.close()
        
        return jsonify({
            'success': True,
            'ml_model': 'active',
            'database': 'active',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'ml_model': 'error',
            'database': 'error',
            'error': str(e)
        }), 500

# ============================================================
# NEW ROUTE: Get Real-Time Traffic Status for Area
# ============================================================
@app.route('/api/traffic-status', methods=['GET'])
def get_traffic_status():
    """
    Get real-time traffic status for a specific area and hour
    
    Query params:
        - area: Area name (Bucal, Parian, Turbina, etc.)
        - hour: Hour of day (0-23)
    
    Returns:
        {
            'success': True,
            'area': 'Bucal',
            'hour': 14,
            'congestion_level': 1.5,  # 0.5-3.0 scale
            'congestion_label': 'Moderate',
            'timestamp': '2025-11-21T14:30:00',
            'data_source': 'realtime' or 'historical' or 'predicted'
        }
    """
    try:
        area = request.args.get('area', '')
        hour = request.args.get('hour', type=int)
        
        if not area or hour is None:
            return jsonify({
                'success': False,
                'error': 'Missing area or hour parameter'
            }), 400
        
        # Map area names to coordinates (adjust these based on your data)
        area_coordinates = {
            'Bucal': {'lat': 14.1894, 'lng': 121.1653},
            'Parian': {'lat': 14.2115, 'lng': 121.1653},
            'Turbina': {'lat': 14.2331, 'lng': 121.1653}
        }
        
        if area not in area_coordinates:
            return jsonify({
                'success': False,
                'error': f'Unknown area: {area}'
            }), 400
        
        coords = area_coordinates[area]
        current_hour = datetime.now().hour
        
        # ============================================================
        # Try to get real-time data for current and recent hours
        # ============================================================
        if hour == current_hour or abs(hour - current_hour) <= 6:
            try:
                realtime_data = traffic_service.get_traffic_flow(
                    coords['lat'], 
                    coords['lng']
                )
                
                if realtime_data.get('success'):
                    congestion_ratio = realtime_data.get('congestion_ratio', 1.0)
                    
                    # Convert 0-2 scale to 0.5-3.0 scale
                    # 0 = light (0.5-1.0)
                    # 1 = moderate (1.0-2.0)
                    # 2 = heavy (2.0-3.0)
                    if congestion_ratio == 0:
                        congestion_level = 0.5 + (random.random() * 0.5)  # 0.5-1.0
                    elif congestion_ratio == 1:
                        congestion_level = 1.0 + (random.random() * 1.0)  # 1.0-2.0
                    else:
                        congestion_level = 2.0 + (random.random() * 1.0)  # 2.0-3.0
                    
                    return jsonify({
                        'success': True,
                        'area': area,
                        'hour': hour,
                        'congestion_level': round(congestion_level, 2),
                        'congestion_label': (
                            'Light' if congestion_level < 1.5 else
                            'Moderate' if congestion_level < 2.5 else
                            'Heavy'
                        ),
                        'timestamp': datetime.now().isoformat(),
                        'data_source': 'realtime'
                    })
            except Exception as e:
                print(f"Real-time API failed: {e}")
                # Fall through to historical/prediction
        
        # ============================================================
        # Use historical patterns + prediction for other hours
        # ============================================================
        
        # Base traffic levels per area (from your DWPH/POSO data)
        base_levels = {
            'Bucal': 1.3,
            'Parian': 1.5,
            'Turbina': 1.2
        }
        
        base_level = base_levels.get(area, 1.0)
        
        # Time-of-day multiplier
        if 6 <= hour <= 9:
            time_mult = 1.6  # Morning rush
        elif 17 <= hour <= 19:
            time_mult = 1.8  # Evening rush
        elif 12 <= hour <= 14:
            time_mult = 1.3  # Lunch
        elif 10 <= hour <= 11:
            time_mult = 1.2  # Mid-morning
        elif 15 <= hour <= 16:
            time_mult = 1.3  # Mid-afternoon
        elif 20 <= hour <= 22:
            time_mult = 1.1  # Evening
        elif hour >= 23 or hour <= 5:
            time_mult = 0.4  # Night
        else:
            time_mult = 0.9  # Off-peak
        
        # Weekend adjustment
        is_weekend = datetime.now().weekday() >= 5
        if is_weekend:
            if 6 <= hour <= 9 or 17 <= hour <= 19:
                time_mult *= 0.7
        
        congestion_level = base_level * time_mult
        
        # Add slight variation
        congestion_level += (random.random() - 0.5) * 0.2
        
        # Clamp to valid range
        congestion_level = max(0.5, min(3.0, congestion_level))
        
        data_source = 'historical' if hour < current_hour else 'predicted'
        
        return jsonify({
            'success': True,
            'area': area,
            'hour': hour,
            'congestion_level': round(congestion_level, 2),
            'congestion_label': (
                'Light' if congestion_level < 1.5 else
                'Moderate' if congestion_level < 2.5 else
                'Heavy'
            ),
            'timestamp': datetime.now().isoformat(),
            'data_source': data_source
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================
# OPTION 2: Query your existing traffic_records table
# ============================================
@app.route('/api/traffic-status-db', methods=['GET'])
def get_traffic_status_from_db():
    """
    Returns actual traffic data from your database.
    Queries the most recent traffic_records for the specified road and hour.
    """
    import psycopg2
    from psycopg2.extras import RealDictCursor
    
    road = request.args.get('road', '')
    hour = request.args.get('hour', type=int)
    
    if not road or hour is None:
        return jsonify({
            "success": False,
            "error": "Missing required parameters: road, hour"
        }), 400
    
    try:
        # Connect to your Supabase database
        conn = psycopg2.connect(
            host="YOUR_SUPABASE_HOST",
            database="postgres",
            user="postgres",
            password="YOUR_PASSWORD",
            port=5432
        )
        
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Query recent traffic records for this road at this hour
        query = """
            SELECT 
                AVG(delay_ratio) as avg_congestion,
                COUNT(*) as record_count
            FROM traffic_records
            WHERE road_segment_id IN (
                SELECT segment_id 
                FROM road_segments 
                WHERE road_name ILIKE %s
            )
            AND EXTRACT(HOUR FROM timestamp) = %s
            AND DATE(timestamp) >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY EXTRACT(HOUR FROM timestamp)
        """
        
        cursor.execute(query, (f'%{road}%', hour))
        result = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if result and result['record_count'] > 0:
            return jsonify({
                "success": True,
                "congestion_level": round(float(result['avg_congestion']), 2),
                "timestamp": datetime.now().isoformat(),
                "data_source": "database",
                "road": road,
                "hour": hour,
                "sample_size": result['record_count']
            })
        else:
            return jsonify({
                "success": False,
                "message": "No recent data available for this road and hour"
            }), 404
            
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

# ============================================================
# NEW ROUTE: Save Simulation to Database
# ============================================================

@app.route('/api/save-simulation', methods=['POST'])
def save_simulation():
    """
    Save a completed simulation to the database
    
    Request body should contain:
    {
        "user_id": 1,  # ID of the user (temporary - will use auth later)
        "simulation_data": {
            "scenario_name": "Roadwork on Bagong Kalsada",
            "description": "Road repair work",
            "disruption_type": "roadwork",
            "area": "Bucal",
            "road_corridor": "Calamba_Pagsanjan",
            "start_datetime": "2025-01-20T08:00:00",
            "end_datetime": "2025-01-20T18:00:00",
            "coordinates": {"lat": 14.1894, "lng": 121.1691}
        },
        "results_data": {
            "summary": {...},
            "hourly_predictions": [...],
            "time_segments": {...}
        }
    }
    """
    try:
        data = request.get_json()
        
        # Extract data
        user_id = data.get('user_id')  # Default to planner1 (id=2)
        simulation_data = data.get('simulation_data', {})
        results_data = data.get('results_data', {})
        
        # Validate required fields
        required_fields = ['scenario_name', 'disruption_type', 'area', 'road_corridor']
        missing_fields = [f for f in required_fields if f not in simulation_data]

        # ✅ Parse datetime strings as-is (they're already in local time)
        start_datetime_str = simulation_data.get('start_datetime')
        end_datetime_str = simulation_data.get('end_datetime')
        
        # Parse without timezone conversion
        start_time = datetime.fromisoformat(start_datetime_str) if start_datetime_str else None
        end_time = datetime.fromisoformat(end_datetime_str) if end_datetime_str else None
        
        print(f"💾 Saving times:")
        print(f"   Start: {start_time}")
        print(f"   End: {end_time}")
        
        if missing_fields:
            return jsonify({
                'success': False,
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
        
        # Save to database
        simulation_id = db.save_simulation_run(
            user_id=user_id,
            simulation_data=simulation_data,
            results_data=results_data
        )
        
        if simulation_id:
            return jsonify({
                'success': True,
                'simulation_id': simulation_id,
                'message': 'Simulation saved successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to save simulation'
            }), 500
            
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================================
# NEW ROUTE: Get User's Simulations
# ============================================================

@app.route('/api/my-simulations', methods=['GET'])
def get_my_simulations():
    """
    Get all simulations for the current user
    
    Query params:
        - user_id: User ID (temporary - will use auth token later)
    """
    try:
        user_id = request.args.get('user_id', type=int)
        
        simulations = db.get_user_simulations(user_id)
        
        return jsonify({
            'success': True,
            'simulations': simulations,
            'count': len(simulations)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================
# NEW ROUTE: Get Simulation Details
# ============================================================

@app.route('/api/simulation/<int:simulation_id>', methods=['GET'])
def get_simulation(simulation_id):
    """Get detailed information about a specific simulation"""
    try:
        simulation = db.get_simulation_by_id(simulation_id)
        
        if simulation:
            # 🔍 ADD THIS LOGGING
            print(f"📊 Fetched simulation {simulation_id}")
            print(f"   - Has hourly_predictions: {simulation.get('hourly_predictions') is not None}")
            print(f"   - Type: {type(simulation.get('hourly_predictions'))}")
            if simulation.get('hourly_predictions'):
                print(f"   - Length: {len(simulation.get('hourly_predictions'))}")
            print(f"   - Has aggregated_view: {simulation.get('aggregated_view') is not None}")
            
            return jsonify({
                'success': True,
                'simulation': simulation
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Simulation not found'
            }), 404
            
    except Exception as e:
        print(f"❌ Error fetching simulation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================
# NEW ROUTE: Publish Simulation
# ============================================================

@app.route('/api/publish-simulation', methods=['POST'])
def publish_simulation():
    """
    Publish a simulation to the public map
    
    Request body:
    {
        "simulation_id": 1,
        "user_id": 2,
        "title": "Roadwork in Bucal",
        "public_description": "Road repair work causing moderate delays"
    }
    """
    try:
        data = request.get_json()
        
        simulation_id = data.get('simulation_id')
        user_id = data.get('user_id')
        title = data.get('title')
        public_description = data.get('public_description')
        
        if not simulation_id:
            return jsonify({
                'success': False,
                'error': 'simulation_id is required'
            }), 400
        
        slug = db.publish_simulation(
            simulation_id=simulation_id,
            published_by_user_id=user_id,
            title=title,
            public_description=public_description
        )
        
        if slug:
            return jsonify({
                'success': True,
                'slug': slug,
                'message': 'Simulation published successfully',
                'public_url': f'/disruptions/{slug}'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to publish simulation'
            }), 500
            
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================================
# NEW ROUTE: Unpublish Simulation
# ============================================================

@app.route('/api/unpublish-simulation', methods=['POST'])
def unpublish_simulation():
    """
    Unpublish a simulation from the public map
    
    Request body:
    {
        "simulation_id": 1,
        "user_id": 2
    }
    """
    try:
        data = request.get_json()
        
        simulation_id = data.get('simulation_id')
        user_id = data.get('user_id')
        
        if not simulation_id:
            return jsonify({
                'success': False,
                'error': 'simulation_id is required'
            }), 400
        
        success = db.unpublish_simulation(simulation_id, user_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Simulation unpublished successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to unpublish simulation'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================
# NEW ROUTE: Get Published Simulations (For Public Map)
# ============================================================

@app.route('/api/published-disruptions', methods=['GET'])
def get_published_disruptions():
    """
    Get all published simulations for the public map
    This replaces the mock data in HomeMapWithSidebar.jsx
    """
    conn = None
    try:
        conn = db._get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get published disruptions with full details
        query = """
            SELECT 
                pr.published_id,
                pr.simulation_id,
                pr.title,
                pr.public_description,
                pr.published_at,
                pr.is_active,
                pr.view_count,
                pr.slug,
                sr.simulation_name,
                sr.description,
                sr.disruption_type,
                sr.disruption_location,
                sr.start_time,
                sr.end_time,
                sr.severity_level,
                sr.total_affected_segments,
                sr.average_delay_ratio,
                ST_Y(sr.disruption_geometry) as latitude,
                ST_X(sr.disruption_geometry) as longitude,
                sr.hourly_predictions::text as hourly_predictions_json,
                u.organization,
                u.full_name
            FROM published_runs pr
            INNER JOIN simulation_runs sr ON pr.simulation_id = sr.simulation_id
            LEFT JOIN users u ON pr.published_by = u.user_id
            WHERE pr.is_active = TRUE
            AND sr.simulation_status != 'deleted'
            ORDER BY pr.published_at DESC
        """
        
        cursor.execute(query)
        simulations = cursor.fetchall()
        
        print(f"\n📊 Fetched {len(simulations)} published disruptions")
        
        # Transform to frontend format
        disruptions = []
        for sim in simulations:
            # Extract coordinates from simulation_data
            # ✅ GET COORDINATES DIRECTLY FROM POSTGIS FUNCTIONS
            latitude = sim.get('latitude')
            longitude = sim.get('longitude')
            road_coordinates = None

            if latitude and longitude:
                print(f"✅ {sim['title']}: lat={latitude}, lng={longitude}")
            else:
                print(f"❌ No coordinates for: {sim['title']}")
                # Use fallback
                location = sim.get('disruption_location', '')
                coords = get_coordinates_for_location(location)
                latitude = coords['lat']
                longitude = coords['lng']

            wkt = sim.get('disruption_geometry_wkt')
            if wkt:
                # Parse "POINT(121.1640 14.2096)"
                import re
                match = re.match(r'POINT\(([0-9.]+)\s+([0-9.]+)\)', wkt)
                if match:
                    longitude = float(match.group(1))
                    latitude = float(match.group(2))
                    print(f"✅ {sim['title']}: lat={latitude}, lng={longitude}")
                else:
                    print(f"❌ Could not parse WKT: {wkt}")
            else:
                print(f"⚠️ No geometry for: {sim['title']}")
            
            simulation_data = sim.get('simulation_data')
            if simulation_data:
                if isinstance(simulation_data, str):
                    import json
                    simulation_data = json.loads(simulation_data)
                
                coords = simulation_data.get('coordinates', {})
                latitude = coords.get('lat')
                longitude = coords.get('lng') or coords.get('lon')
                
                # Get road coordinates if available (for drawing lines)
                road_coordinates = simulation_data.get('road_coordinates_json')
            
            # Fallback to location-based coordinates if not in simulation_data
            if not latitude or not longitude:
                location = sim.get('disruption_location', '')
                coords = get_coordinates_for_location(location)
                latitude = coords['lat']
                longitude = coords['lng']
            
            # Extract hourly predictions from results_data
            hourly_predictions = None
            results_data = sim.get('results_data')
            if results_data:
                if isinstance(results_data, str):
                    import json
                    results_data = json.loads(results_data)
                
                # Get hourly predictions for active disruptions
                hourly_predictions = results_data.get('hourly_predictions')
            
            # Calculate expected delay
            avg_delay_ratio = float(sim.get('average_delay_ratio') or 1.0)
            expected_delay = round(avg_delay_ratio * 10)
            
            # Determine congestion level
            severity_level = sim.get('severity_level', '').lower()
            if severity_level in ['heavy', 'high']:
                congestion_level = 'Heavy'
            elif severity_level in ['moderate', 'medium']:
                congestion_level = 'Moderate'
            elif severity_level in ['light', 'low']:
                congestion_level = 'Light'
            else:
                if avg_delay_ratio >= 2.0:
                    congestion_level = 'Heavy'
                elif avg_delay_ratio >= 1.5:
                    congestion_level = 'Moderate'
                else:
                    congestion_level = 'Light'
            
            disruption = {
                'id': sim['published_id'],
                'simulation_id': sim['simulation_id'],
                'title': sim['title'] or sim['simulation_name'],
                'description': sim['public_description'] or sim.get('description', ''),
                'location': sim['disruption_location'] or 'Calamba City',
                'type': sim['disruption_type'] or 'general',
                'status': 'Active',
                'start_date': sim['start_time'].isoformat() if sim.get('start_time') else None,
                'end_date': sim['end_time'].isoformat() if sim.get('end_time') else None,
                'severity_level': severity_level,
                'congestion_level': congestion_level,
                'avg_severity': avg_delay_ratio,
                'expected_delay': expected_delay,
                'published_at': sim['published_at'].isoformat() if sim.get('published_at') else None,
                'slug': sim.get('slug', ''),
                'view_count': sim.get('view_count', 0),
                'organization': sim.get('organization', 'DPWH'),
                'latitude': latitude,
                'longitude': longitude,
                'road_coordinates': road_coordinates,  # For drawing road lines
                'hourly_predictions': hourly_predictions  # For detailed visualization
            }
            
            disruptions.append(disruption)
        
        print(f"✅ Returning {len(disruptions)} disruptions with coordinates")
        for d in disruptions[:3]:
            print(f"   - {d['title']}: ({d['latitude']}, {d['longitude']})")
        
        return jsonify({
            'success': True,
            'disruptions': disruptions,
            'count': len(disruptions)
        })
        
    except Exception as e:
        import traceback
        print(f"\n❌ Error in get_published_disruptions:")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
    finally:
        if conn:
            conn.close()

'''
def get_coordinates_for_location(location):
    """
    Get coordinates for a location string
    This is a fallback when coordinates aren't in simulation_data
    """
    location_lower = location.lower()
    
    # Area to coordinates mapping
    area_coordinates = {
        'bucal': {'lat': 14.1894, 'lng': 121.1653},
        'parian': {'lat': 14.2115, 'lng': 121.1653},
        'turbina': {'lat': 14.2331, 'lng': 121.1653},
        'bagong kalsada': {'lat': 14.2050, 'lng': 121.1620},
        'real': {'lat': 14.2150, 'lng': 121.1580},
        'halang': {'lat': 14.2200, 'lng': 121.1700},
        'crossing': {'lat': 14.2120, 'lng': 121.1640},
        'pansol': {'lat': 14.1980, 'lng': 121.1750},
        'makiling': {'lat': 14.2450, 'lng': 121.1700},
        'milagrosa': {'lat': 14.2280, 'lng': 121.1650},
        'checkpoint': {'lat': 14.2180, 'lng': 121.1620}
    }
    
    # Try to match area
    for area, coords in area_coordinates.items():
        if area in location_lower:
            print(f"📍 Matched '{location}' to {area}: ({coords['lat']}, {coords['lng']})")
            return coords
    
    # Default to Calamba center
    print(f"📍 Using default coordinates for '{location}'")
    return {'lat': 14.2096, 'lng': 121.1640}
'''

def get_coordinates_for_location(location):
    """Get coordinates for a location string"""
    location_lower = location.lower()
    
    # UPDATED coordinates for Calamba City areas
    area_coordinates = {
        'bucal': {'lat': 14.2118, 'lng': 121.1645},
        'parian': {'lat': 14.2115, 'lng': 121.1635},
        'turbina': {'lat': 14.2186, 'lng': 121.1580},
        'bagong kalsada': {'lat': 14.2000, 'lng': 121.1755},
        'real': {'lat': 14.2207, 'lng': 121.1567},
        'halang': {'lat': 14.1985, 'lng': 121.1780},
        'crossing': {'lat': 14.2096, 'lng': 121.1640},
        'pansol': {'lat': 14.1947, 'lng': 121.1800},
        'makiling': {'lat': 14.2331, 'lng': 121.1653},
        'milagrosa': {'lat': 14.2280, 'lng': 121.1605},
        'checkpoint': {'lat': 14.2148, 'lng': 121.1610},
        'calamba': {'lat': 14.2096, 'lng': 121.1640}
    }
    
    for area, coords in area_coordinates.items():
        if area in location_lower:
            return coords
    
    return {'lat': 14.2096, 'lng': 121.1640}

# Alias route for frontend compatibility
@app.route('/api/published-simulations', methods=['GET'])
def get_published_simulations_alias():
    """
    Alias for /api/published-disruptions for frontend compatibility
    """
    return get_published_disruptions()

# ============================================================
# NEW ROUTE: Delete Simulation
# ============================================================

@app.route('/api/delete-simulation/<int:simulation_id>', methods=['DELETE'])
def delete_simulation(simulation_id):
    """
    Delete (soft delete) a simulation
    
    Query params:
        - user_id: User ID (temporary - will use auth later)
    """
    try:
        user_id = request.args.get('user_id', type=int)
        
        success = db.delete_simulation(simulation_id, user_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Simulation deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to delete simulation or not authorized'
            }), 403
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================
# UPDATED ROUTE: Simulate Disruption (Save After Simulation)
# ============================================================


# ============================================================
# NEW ROUTE: Get Real-Time Traffic for Location
# ============================================================

@app.route('/api/realtime-traffic', methods=['POST'])
def get_realtime_traffic():
    """
    Get current real-time traffic data for a specific location
    """
    try:
        data = request.get_json()
        lat = data.get('lat')
        lng = data.get('lng')
        
        if not lat or not lng:
            return jsonify({
                'success': False,
                'error': 'Missing lat/lng'
            }), 400
        
        # Fetch real-time traffic
        traffic_data = traffic_service.get_traffic_flow(lat, lng)
        
        return jsonify(traffic_data)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# NEW ROUTE: Get Traffic for Road Segments
# ============================================================

@app.route('/api/realtime-road-traffic', methods=['POST'])
def get_road_traffic():
    """
    Get real-time traffic for multiple points along a road
    """
    try:
        data = request.get_json()
        coordinates = data.get('coordinates', [])
        
        if not coordinates:
            return jsonify({
                'success': False,
                'error': 'No coordinates provided'
            }), 400
        
        # Sample coordinates (e.g., every 5th point to avoid API limits)
        sampled_coords = coordinates[::5]  # Every 5th point
        
        # Get traffic for sampled points
        traffic_results = traffic_service.get_multiple_segments(sampled_coords)
        
        # Calculate average congestion
        valid_results = [r for r in traffic_results if r.get('success')]
        if valid_results:
            avg_congestion = sum(r.get('congestion_ratio', 0) for r in valid_results) / len(valid_results)
            avg_speed = sum(r.get('current_speed', 0) for r in valid_results) / len(valid_results)
        else:
            avg_congestion = 0
            avg_speed = 0
        
        return jsonify({
            'success': True,
            'segments': traffic_results,
            'summary': {
                'avg_congestion_ratio': avg_congestion,
                'avg_current_speed': avg_speed,
                'total_segments_checked': len(sampled_coords),
                'successful_checks': len(valid_results)
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# UPDATED: Simulation with Real-Time Data Integration
# ============================================================

@app.route('/api/simulate-disruption-realtime', methods=['POST'])
def simulate_disruption_realtime():
    """
    Smart simulation that uses real-time data ONLY for same-day disruptions (0-6 hours ahead)
    Future disruptions use historical patterns only
    """
    try:
        data = request.get_json()
        
        # ✅ VALIDATE REQUIRED FIELDS FIRST
        required_fields = ['area', 'road_corridor', 'disruption_type', 'start_date', 'start_time', 'end_date', 'end_time']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                'success': False,
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
        
        # Extract basic parameters
        area = str(data.get('area', 'Unknown'))
        road_corridor = str(data.get('road_corridor', 'Unknown'))
        disruption_type = str(data.get('disruption_type'))
        
        # ✅ PARSE DATETIME - This automatically handles string-to-datetime conversion
        try:
            start_datetime = datetime.strptime(
                f"{data['start_date']} {data['start_time']}", 
                "%Y-%m-%d %H:%M"
            )
            end_datetime = datetime.strptime(
                f"{data['end_date']} {data['end_time']}", 
                "%Y-%m-%d %H:%M"
            )
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': f'Invalid date/time format: {str(e)}'
            }), 400

        # ✅ VALIDATE DATES
        if end_datetime <= start_datetime:
            return jsonify({
                'success': False,
                'error': 'End date/time must be after start date/time'
            }), 400
        
        duration_hours = (end_datetime - start_datetime).total_seconds() / 3600
        
        if duration_hours > 720:  # 30 days
            return jsonify({
                'success': False,
                'error': 'Disruption duration cannot exceed 30 days'
            }), 400
        
        if duration_hours < 1:
            return jsonify({
                'success': False,
                'error': 'Disruption duration must be at least 1 hour'
            }), 400
        
        road_info = data.get('road_info', {})
        coordinates = data.get('coordinates', {})
        
        # ✅ VALIDATE COORDINATES WITH TYPE CONVERSION
        try:
            lat = float(coordinates.get('lat', 0))
            lng = float(coordinates.get('lng', 0))
            
            if lat == 0 or lng == 0:
                raise ValueError("Invalid coordinates")
                
            coordinates = {'lat': lat, 'lng': lng}
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'Invalid or missing location coordinates'
            }), 400
        
        # ✅ CONVERT ROAD INFO NUMERIC VALUES
        try:
            road_info = {
                'lanes': int(road_info.get('lanes', 2)),
                'length_km': float(road_info.get('length_km', 1.0)),
                'width_meters': float(road_info.get('width_meters', 7.0)),
                'max_speed': int(road_info.get('max_speed', 40)),
                'total_capacity': int(road_info.get('total_capacity', 1800)),
                'free_flow_time_minutes': float(road_info.get('free_flow_time_minutes', 10)),
                'disruption_factors': road_info.get('disruption_factors', {}),
                'road_type': str(road_info.get('road_type', 'local'))
            }
        except (ValueError, TypeError) as e:
            return jsonify({
                'success': False,
                'error': f'Invalid road info data: {str(e)}'
            }), 400
        
        # ============================================================
        # ✅ SMART DECISION: Should we use real-time data?
        # ============================================================
        
        now = datetime.now()
        today = now.date()
        disruption_start_date = start_datetime.date()
        
        # Calculate hours until disruption starts
        hours_until_disruption = (start_datetime - now).total_seconds() / 3600
        
        # ✅ USE REAL-TIME ONLY IF:
        # 1. Disruption starts today
        # 2. Within next 6 hours (or already started)
        use_realtime = (
            disruption_start_date == today and 
            hours_until_disruption <= 6 and 
            hours_until_disruption >= -24  # Allow up to 24h past start (ongoing disruption)
        )
        
        realtime_data = None
        realtime_speed_factor = None
        current_congestion = 0
        
        if use_realtime:
            print("\n" + "="*60)
            print("🌐 FETCHING REAL-TIME TRAFFIC DATA")
            print("="*60)
            print(f"📍 Location: {coordinates['lat']}, {coordinates['lng']}")
            print(f"⏰ Current Time: {now.strftime('%Y-%m-%d %H:%M')}")
            print(f"🚧 Disruption Start: {start_datetime.strftime('%Y-%m-%d %H:%M')}")
            print(f"⏱️  Hours Until Start: {hours_until_disruption:.1f}h")
            print(f"✅ USING REAL-TIME (disruption happening soon)")
            
            # Fetch real-time traffic
            realtime_data = traffic_service.get_traffic_flow(
                coordinates['lat'],
                coordinates['lng']
            )
            
            if realtime_data.get('success'):
                current_speed = float(realtime_data.get('current_speed', 40))
                free_flow_speed = float(realtime_data.get('free_flow_speed', 40))
                current_congestion = int(realtime_data.get('congestion_ratio', 0))
                
                if free_flow_speed > 0:
                    realtime_speed_factor = current_speed / free_flow_speed
                
                print(f"✅ Real-time API Success")
                print(f"🚗 Current Speed: {current_speed} km/h")
                print(f"🏁 Free Flow Speed: {free_flow_speed} km/h")
                print(f"📊 Speed Factor: {realtime_speed_factor:.2f}")
                print(f"📊 Congestion Level: {current_congestion} (0=light, 1=moderate, 2=heavy)")
            else:
                print(f"❌ Real-time API Failed: {realtime_data.get('error')}")
                use_realtime = False  # Fall back to historical
            
            print("="*60 + "\n")
            
        else:
            print("\n" + "="*60)
            print("📅 FUTURE DISRUPTION - USING HISTORICAL PATTERNS")
            print("="*60)
            print(f"⏰ Current Time: {now.strftime('%Y-%m-%d %H:%M')}")
            print(f"🚧 Disruption Start: {start_datetime.strftime('%Y-%m-%d %H:%M')}")
            
            if disruption_start_date > today:
                days_away = (disruption_start_date - today).days
                print(f"📆 Disruption is {days_away} day(s) away")
            else:
                print(f"⏱️  Disruption is {hours_until_disruption:.1f} hours away")
            
            print(f"✅ Using ML model trained on historical data")
            print(f"📊 Real-time data not applicable for future predictions")
            print("="*60 + "\n")
        
        # ============================================================
        # Generate Hourly Predictions
        # ============================================================
        
        hourly_predictions = []
        current_datetime = start_datetime
        
        while current_datetime <= end_datetime:
            # ✅ ENSURE ALL VALUES ARE PROPER TYPES
            hour_input = {
                'date': current_datetime.strftime('%Y-%m-%d'),
                'hour': int(current_datetime.hour),  # ✅ ENSURE INT
                'area': str(area),
                'road_corridor': str(road_corridor),
                'has_disruption': 1,
                'disruption_type': str(disruption_type),
                'total_volume': int(data.get('total_volume', 0)),
                'has_real_status': 0
            }
            
            # Make prediction using ML model
            prediction = predictor.predict(hour_input)
            
            # ✅ ONLY apply real-time factor to predictions within next 6 hours
            hours_until_this_prediction = (current_datetime - now).total_seconds() / 3600
            apply_realtime_to_this_hour = (
                use_realtime and 
                -1 <= hours_until_this_prediction <= 6  # Current hour to 6 hours ahead
            )
            
            # Calculate delay with optional real-time adjustment
            delay_info = predictor.estimate_delay(
                severity=float(prediction['severity']),
                base_travel_time_minutes=float(road_info['free_flow_time_minutes']),
                road_length_km=float(road_info['length_km']),
                impact_factor=float(road_info['disruption_factors'].get(disruption_type, 0.6)),
                realtime_speed_factor=realtime_speed_factor if apply_realtime_to_this_hour else None
            )
            
            hourly_predictions.append({
                'datetime': current_datetime.strftime('%Y-%m-%d %H:%M'),
                'date': current_datetime.strftime('%Y-%m-%d'),
                'hour': int(current_datetime.hour),  # ✅ ENSURE INT
                'day_of_week': current_datetime.strftime('%A'),
                'severity': round(float(prediction['severity']), 2),
                'severity_label': str(prediction['severity_label']),
                'confidence': round(float(prediction['confidence']), 2),
                'delay_info': delay_info,
                'realtime_adjusted': bool(delay_info.get('realtime_adjusted', False)),
                'probabilities': {
                    k: round(float(v), 2) for k, v in prediction['probabilities'].items()
                }
            })
            
            current_datetime += timedelta(hours=1)
        
        # ============================================================
        # Smart Aggregation for Map Display
        # ============================================================

        def aggregate_predictions_smart(hourly_predictions, start_datetime, end_datetime):
            """
            Aggregate predictions on a day-by-day basis
            Always returns daily granularity regardless of duration
            """
            # Always show day-by-day breakdown
            daily_aggregates = []
            current_date = start_datetime.date()
            end_date = end_datetime.date()
            
            while current_date <= end_date:
                # Get all predictions for this day
                day_predictions = [
                    p for p in hourly_predictions 
                    if datetime.strptime(p['datetime'], '%Y-%m-%d %H:%M').date() == current_date
                ]
                
                if day_predictions:
                    # Calculate daily averages
                    avg_severity = sum(p['severity'] for p in day_predictions) / len(day_predictions)
                    avg_delay = sum(p['delay_info']['additional_delay_min'] for p in day_predictions) / len(day_predictions)
                    
                    # Determine dominant severity
                    severity_counts = {
                        'Light': sum(1 for p in day_predictions if p['severity'] < 0.5),
                        'Moderate': sum(1 for p in day_predictions if 0.5 <= p['severity'] < 1.5),
                        'Heavy': sum(1 for p in day_predictions if p['severity'] >= 1.5)
                    }
                    dominant_severity = max(severity_counts, key=severity_counts.get)
                    
                    # ✅ FIX: Find ALL peak hours and calculate peak delay
                    max_severity = max(p['severity'] for p in day_predictions)
                    # Consider hours within 0.1 of max severity as "peak hours"
                    peak_threshold = max_severity - 0.1
                    peak_hour_predictions = [p for p in day_predictions if p['severity'] >= peak_threshold]
                    
                    # Get the peak hours (sorted)
                    peak_hours_list = sorted([p['hour'] for p in peak_hour_predictions])
                    
                    # Calculate average delay during peak hours
                    peak_delay = sum(p['delay_info']['additional_delay_min'] for p in peak_hour_predictions) / len(peak_hour_predictions) if peak_hour_predictions else 0
                    
                    # Format peak hour display
                    if len(peak_hours_list) == 1:
                        peak_hour_display = peak_hours_list[0]
                    elif len(peak_hours_list) <= 3:
                        # Show as range or list for 2-3 hours
                        peak_hour_display = peak_hours_list[0] if len(peak_hours_list) == 1 else f"{peak_hours_list[0]}-{peak_hours_list[-1]}"
                    else:
                        # Multiple peak hours - show first and last
                        peak_hour_display = f"{peak_hours_list[0]}-{peak_hours_list[-1]}"
                    
                    daily_aggregates.append({
                        'date': current_date.strftime('%Y-%m-%d'),
                        'day_name': current_date.strftime('%A'),
                        'avg_severity': round(avg_severity, 2),
                        'avg_severity_label': dominant_severity,
                        'avg_delay_min': round(avg_delay),
                        'hour_count': len(day_predictions),
                        'severity_breakdown': severity_counts,
                        'peak_hour': peak_hour_display,  # ✅ Now shows formatted hour(s)
                        'peak_hours': peak_hours_list,  # ✅ Array of all peak hours
                        'peak_severity': round(max_severity, 2),
                        'peak_delay': round(peak_delay),  # ✅ Average delay during peak hours
                        'avg_peak_delay': round(peak_delay)  # ✅ Alternative field name for compatibility
                    })
                
                current_date += timedelta(days=1)
            
            return {
                'granularity': 'daily',
                'display_label': 'Day-by-Day View',
                'map_data': daily_aggregates
            }

        # ✅ Generate aggregated view
        aggregated_view = aggregate_predictions_smart(hourly_predictions, start_datetime, end_datetime)
        
        # ============================================================
        # Calculate Summary Statistics
        # ============================================================
        
        total_hours = len(hourly_predictions)

        # ✅ SAFETY CHECK - Prevent division by zero
        if total_hours == 0:
            return jsonify({
                'success': False,
                'error': 'No predictions generated. Check start and end dates.'
            }), 400

        light_hours = sum(1 for p in hourly_predictions if p['severity'] < 0.5)
        moderate_hours = sum(1 for p in hourly_predictions if 0.5 <= p['severity'] < 1.5)
        heavy_hours = sum(1 for p in hourly_predictions if p['severity'] >= 1.5)
        
        # ✅ SAFE DIVISION
        avg_severity = sum(p['severity'] for p in hourly_predictions) / max(total_hours, 1)
        avg_delay = sum(p['delay_info']['additional_delay_min'] for p in hourly_predictions) / max(total_hours, 1)

        # Time segment breakdown
        time_segments = {
            'morning': {'light': 0, 'moderate': 0, 'heavy': 0},
            'afternoon': {'light': 0, 'moderate': 0, 'heavy': 0},
            'night': {'light': 0, 'moderate': 0, 'heavy': 0}
        }
        
        for pred in hourly_predictions:
            hour = int(pred['hour'])  # ✅ ENSURE INT
            if pred['severity'] < 0.5:
                sev_label = 'light'
            elif pred['severity'] < 1.5:
                sev_label = 'moderate'
            else:
                sev_label = 'heavy'
            
            if 6 <= hour <= 11:
                time_segments['morning'][sev_label] += 1
            elif 12 <= hour <= 17:
                time_segments['afternoon'][sev_label] += 1
            else:
                time_segments['night'][sev_label] += 1
        
        simulation_id = f"sim_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # ============================================================
        # Calculate Affected Road Segments with Individual Severities
        # ============================================================
        
        affected_segments = []
        
        # Define impact zones with distance-based severity multipliers
        impact_zones = [
            {'min_dist': 0, 'max_dist': 150, 'multiplier': 1.0, 'label': 'Critical Impact'},
            {'min_dist': 150, 'max_dist': 300, 'multiplier': 0.75, 'label': 'High Impact'},
            {'min_dist': 300, 'max_dist': 450, 'multiplier': 0.50, 'label': 'Moderate Impact'},
            {'min_dist': 450, 'max_dist': 600, 'multiplier': 0.30, 'label': 'Low Impact'},
        ]
        
        # Road type importance factors (higher capacity = more affected by disruption)
        road_type_factors = {
            'motorway': 1.2,
            'trunk': 1.15,
            'primary': 1.1,
            'secondary': 1.0,
            'tertiary': 0.9,
            'residential': 0.7,
            'service': 0.5,
        }
        
        # Create segment data for main road
        main_segment = {
            'segment_id': 'main',
            'road_name': road_corridor,
            'road_type': road_info.get('road_type', 'secondary'),
            'distance_m': 0,
            'impact_zone': 'Critical Impact',
            'avg_severity': round(avg_severity, 2),
            'severity_label': 'Light' if avg_severity < 0.5 else ('Moderate' if avg_severity < 1.5 else 'Heavy'),
            'avg_delay_min': round(avg_delay),
            'hourly_severities': [
                {
                    'hour': p['hour'],
                    'datetime': p['datetime'],
                    'severity': p['severity'],
                    'severity_label': p['severity_label'],
                    'delay_min': p['delay_info']['additional_delay_min']
                }
                for p in hourly_predictions
            ]
        }
        affected_segments.append(main_segment)
        
        # Generate nearby road segment predictions
        # These will be used by frontend to color nearby roads accurately
        sample_nearby_roads = [
            {'id': 'nearby_1', 'name': 'Adjacent Road 1', 'type': 'secondary', 'distance': 100},
            {'id': 'nearby_2', 'name': 'Adjacent Road 2', 'type': 'tertiary', 'distance': 200},
            {'id': 'nearby_3', 'name': 'Adjacent Road 3', 'type': 'secondary', 'distance': 350},
            {'id': 'nearby_4', 'name': 'Adjacent Road 4', 'type': 'residential', 'distance': 500},
        ]
        
        for nearby in sample_nearby_roads:
            # Find applicable impact zone
            zone = next(
                (z for z in impact_zones if z['min_dist'] <= nearby['distance'] < z['max_dist']),
                impact_zones[-1]
            )
            
            # Calculate segment-specific severity
            road_factor = road_type_factors.get(nearby['type'], 0.8)
            segment_severity = avg_severity * zone['multiplier'] * road_factor
            segment_delay = avg_delay * zone['multiplier'] * road_factor
            
            segment = {
                'segment_id': nearby['id'],
                'road_name': nearby['name'],
                'road_type': nearby['type'],
                'distance_m': nearby['distance'],
                'impact_zone': zone['label'],
                'impact_multiplier': zone['multiplier'],
                'road_factor': road_factor,
                'avg_severity': round(segment_severity, 2),
                'severity_label': 'Light' if segment_severity < 0.5 else ('Moderate' if segment_severity < 1.5 else 'Heavy'),
                'avg_delay_min': round(segment_delay),
            }
            affected_segments.append(segment)

        # ============================================================
        # Return Response
        # ============================================================
        
        return jsonify({
            'success': True,
            'simulation_id': simulation_id,
            'realtime_integration': {
                'enabled': use_realtime and realtime_data and realtime_data.get('success', False),
                'applicable': use_realtime,
                'reason': (
                    'Same-day disruption - adjusted for current traffic' if use_realtime 
                    else f'Future disruption ({(disruption_start_date - today).days} days away) - using historical patterns'
                ),
                'current_speed': round(realtime_data.get('current_speed', 0), 1) if realtime_data else None,
                'free_flow_speed': round(realtime_data.get('free_flow_speed', 0), 1) if realtime_data else None,
                'speed_factor': round(realtime_speed_factor, 2) if realtime_speed_factor else None,
                'current_congestion': current_congestion,
                'timestamp': realtime_data.get('timestamp') if realtime_data else None,
                'hours_adjusted': sum(1 for p in hourly_predictions if p['realtime_adjusted'])
            },
            'input': {
                'area': area,
                'road_corridor': road_corridor,
                'disruption_type': disruption_type,
                'start': start_datetime.strftime('%Y-%m-%d %H:%M'),
                'end': end_datetime.strftime('%Y-%m-%d %H:%M'),
                'description': data.get('description', ''),
                'coordinates': coordinates,
                'road_info': road_info
            },
            'summary': {
                'total_hours': total_hours,
                'duration_days': round((end_datetime - start_datetime).total_seconds() / 86400, 1),
                'light_hours': light_hours,
                'moderate_hours': moderate_hours,
                'heavy_hours': heavy_hours,
                'light_percentage': round(light_hours / total_hours * 100, 1),
                'moderate_percentage': round(moderate_hours / total_hours * 100, 1),
                'heavy_percentage': round(heavy_hours / total_hours * 100, 1),
                'avg_severity': round(avg_severity, 1),
                'avg_severity_label': 'Light' if avg_severity < 0.5 else ('Moderate' if avg_severity < 1.5 else 'Heavy'),
                'avg_delay_minutes': round(avg_delay),
                'total_delay_hours': round(sum(p['delay_info']['additional_delay_min'] for p in hourly_predictions) / 60, 1)
            },
            'hourly_predictions': hourly_predictions,
            'time_segments': time_segments,
            'aggregated_view': aggregated_view,
            'has_multiple_days': (end_datetime - start_datetime).days > 1,
            'affected_segments': affected_segments,
            'impact_zones': impact_zones,
            'road_info': road_info  # ✅ ADD THIS - needed for frontend
        })

    except Exception as e:
        import traceback
        print("\n❌ ERROR in simulate_disruption_realtime:")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc() if app.debug else None
        }), 500
# ============================================================
# ROUTE 1: Home Page
# ============================================================

@app.route('/')
def home():
    """Render the main page"""
    return render_template('index.html')


# ============================================================
# ROUTE 3: Get Road Info (KEEP THIS - It's still useful!)
# ============================================================

@app.route('/api/get-road-info', methods=['POST'])
def get_road_info():
    """
    Simple coordinate to area mapping (fallback if OSM fails)
    KEEP THIS - Frontend uses it as backup
    """
    data = request.get_json()
    lat = data.get('lat')
    lon = data.get('lon')
    
    # Define approximate bounding boxes for each area
    areas = {
        'Bucal': {
            'lat_range': (14.18, 14.20),
            'lon_range': (121.16, 121.18),
            'road_corridor': 'Calamba_Pagsanjan',
            'road_name': 'Calamba-Pagsanjan Road'
        },
        'Parian': {
            'lat_range': (14.21, 14.22),
            'lon_range': (121.14, 121.16),
            'road_corridor': 'Maharlika_Parian',
            'road_name': 'Maharlika Highway (Parian Section)'
        },
        'Turbina': {
            'lat_range': (14.18, 14.19),
            'lon_range': (121.13, 121.15),
            'road_corridor': 'Maharlika_Turbina',
            'road_name': 'Maharlika Highway (Turbina Section)'
        }
    }
    
    for area_name, area_info in areas.items():
        lat_min, lat_max = area_info['lat_range']
        lon_min, lon_max = area_info['lon_range']
        
        if lat_min <= lat <= lat_max and lon_min <= lon <= lon_max:
            return jsonify({
                'success': True,
                'area': area_name,
                'road_corridor': area_info['road_corridor'],
                'road_name': area_info['road_name'],
                'coordinates': {'lat': lat, 'lon': lon}
            })
    
    return jsonify({
        'success': False,
        'message': 'Location not in covered area',
        'coordinates': {'lat': lat, 'lon': lon}
    })

# ============================================================
# ROUTE 4: NEW - Process Road Info from OSM
# ============================================================

@app.route('/api/process-road-info', methods=['POST'])
def process_road_info():
    """
    NEW ENDPOINT - Process OSM road data and calculate capacities
    """
    try:
        data = request.get_json()
        
        # Extract road parameters
        lanes = data.get('lanes', 2)
        length_km = float(data.get('length_km', 1.0))
        width_meters = float(data.get('width_meters', 7.0))
        max_speed = int(data.get('max_speed', 40))
        road_type = data.get('road_type', 'tertiary')
        
        # Calculate road capacity (vehicles per hour)
        lane_capacity = calculate_lane_capacity(road_type, max_speed)
        total_capacity = lane_capacity * lanes
        
        # Calculate free-flow travel time (minutes)
        free_flow_time = (length_km / max_speed) * 60
        
        # Estimate congestion thresholds
        capacity_thresholds = {
            'light': total_capacity * 0.4,
            'moderate': total_capacity * 0.7,
            'heavy': total_capacity * 1.0
        }
        
        # Calculate disruption impact factors
        disruption_factors = calculate_disruption_factors(
            lanes=lanes,
            length_km=length_km,
            road_type=road_type
        )
        
        return jsonify({
            'success': True,
            'road_info': {
                **data,  # Include all original data from OSM
                'lane_capacity': lane_capacity,
                'total_capacity': total_capacity,
                'free_flow_time_minutes': round(free_flow_time, 2),
                'capacity_thresholds': capacity_thresholds,
                'disruption_factors': disruption_factors,
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_lane_capacity(road_type, max_speed):
    """
    Calculate per-lane capacity using Highway Capacity Manual (HCM) 2016 methodology
    
    Reference: Transportation Research Board. Highway Capacity Manual 6th Edition (2016)
    Basic freeway sections: 2,400 pc/h/ln under ideal conditions
    Adjustment factors applied for speed, road type, and local conditions
    """
    # HCM 2016 base capacity for ideal conditions (passenger cars per hour per lane)
    ideal_capacity = 2400
    
    # Speed adjustment factor (from HCM 2016 Exhibit 11-7)
    # Free-flow speed categories and their capacity adjustments
    if max_speed >= 100:  # >= 100 km/h (Freeways)
        speed_factor = 1.00
    elif max_speed >= 80:  # 80-100 km/h (Expressways)
        speed_factor = 0.95
    elif max_speed >= 60:  # 60-80 km/h (Primary arterials)
        speed_factor = 0.85
    elif max_speed >= 40:  # 40-60 km/h (Secondary roads)
        speed_factor = 0.70
    else:  # < 40 km/h (Local/residential)
        speed_factor = 0.55
    
    # Road type adjustment (from HCM principles and Philippine context)
    road_type_factors = {
        'motorway': 1.00,    # Controlled access, ideal
        'trunk': 0.92,       # Major highways, some access points
        'primary': 0.83,     # Arterials with traffic signals
        'secondary': 0.75,   # More intersections and friction
        'tertiary': 0.65,    # Local collectors, high friction
        'residential': 0.50, # Frequent stops, parking, pedestrians
    }
    
    type_factor = road_type_factors.get(road_type, 0.70)
    
    # Philippine context adjustment factor (based on local studies)
    # Reference: JICA Study on Metro Manila Urban Transportation (2014)
    # Accounts for: mixed traffic, weak lane discipline, roadside friction
    ph_adjustment = 0.85
    capacity = ideal_capacity * speed_factor * type_factor * ph_adjustment
    return int(capacity)


def calculate_disruption_factors(lanes, length_km, road_type):
    """
    Calculate disruption impact using empirical models from traffic engineering research
    
    References:
    1. FHWA Work Zone Road User Costs (2011) - Lane closure impacts
    2. Kwon et al. (2006) "Estimating Time-Varying Capacity" - Incident impacts
    3. Cambridge Systematics (2005) - Event-based congestion studies
    """
    
    # Base impact ratios from FHWA Work Zone studies
    # These represent typical capacity reduction percentages
    base_impacts = {
        'roadwork': 0.55,   # 45% capacity reduction (FHWA 2011: 40-60% range)
        'accident': 0.35,   # 65% capacity reduction (Kwon 2006: 50-80% range)
        'event': 0.70,      # 30% capacity reduction (Cambridge 2005: 25-40% range)
        'weather': 0.75,    # 25% capacity reduction (HCM 2016: 10-40% depending on severity)
    }
    
    # Lane closure impact using the "n-1 lane model"
    # Reference: Banks (1991) "The Two-Capacity Phenomenon"
    # Capacity with n-1 lanes ≠ (n-1)/n of original capacity
    if lanes == 1:
        lane_factor = 1.00  # Single lane closure = total blockage
    elif lanes == 2:
        lane_factor = 0.60  # 2→1 loses ~40% capacity (not 50%)
    elif lanes == 3:
        lane_factor = 0.73  # 3→2 loses ~27% (better than linear)
    elif lanes == 4:
        lane_factor = 0.80  # 4→3 loses ~20%
    else:
        # For 5+ lanes, use logarithmic decay
        lane_factor = 1.0 - (1.0 / (lanes ** 0.5))
    
    # Length impact using queueing theory principles
    # Longer disruptions → longer queue buildup → worse delay
    # Reference: Newell (1982) "Applications of Queueing Theory"
    if length_km < 0.5:
        length_factor = 0.90  # Short disruption, quick recovery
    elif length_km < 1.0:
        length_factor = 1.00  # Baseline
    elif length_km < 2.0:
        length_factor = 1.10  # Extended queue formation
    else:
        # Beyond 2km, use square root relationship
        length_factor = 1.10 + (0.05 * (length_km - 2.0) ** 0.5)
        length_factor = min(length_factor, 1.30)  # Cap at 30% increase
    
    # Road importance multiplier
    # Higher-order roads have more regional impact
    importance = {
        'motorway': 1.25,   # Critical network links
        'trunk': 1.15,      # Major arterials
        'primary': 1.00,    # Baseline
        'secondary': 0.90,  # Local impact
        'tertiary': 0.80,   # Minimal network effect
    }.get(road_type, 0.90)
    
    return {
        disruption: round(impact * lane_factor * length_factor * importance, 3)
        for disruption, impact in base_impacts.items()
    }

# ============================================================
# ROUTE 5: Single Prediction (Keep for testing)
# ============================================================

@app.route('/api/predict', methods=['POST'])
def predict_single():
    """Single prediction endpoint - KEEP THIS"""
    try:
        data = request.get_json()
        
        required_fields = ['date', 'hour', 'area', 'road_corridor']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }), 400
        
        result = predictor.predict(data)

        # Calculate delay estimate
        delay_info = predictor.estimate_delay(
            severity=result['severity'],
            base_travel_time_minutes=10,
            road_length_km=5
        )
        result['delay_info'] = delay_info
        
        return jsonify({
            'success': True,
            'prediction': result,
            'input': data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================
# ROUTE 6: UPDATED - Simulate Disruption with Road Info
# ============================================================

@app.route('/api/simulate-disruption', methods=['POST'])
def simulate_disruption():
    """
    UPDATED VERSION - Now accepts road_info from OSM
    """
    try:
        data = request.get_json()
        
        # Extract basic disruption parameters
        area = data.get('area', 'Unknown')
        road_corridor = data.get('road_corridor', 'Unknown')
        disruption_type = data.get('disruption_type')
        start_datetime = datetime.strptime(
            f"{data['start_date']} {data['start_time']}", 
            "%Y-%m-%d %H:%M"
        )
        end_datetime = datetime.strptime(
            f"{data['end_date']} {data['end_time']}", 
            "%Y-%m-%d %H:%M"
        )
        
        # NEW: Get road information (from OSM)
        road_info = data.get('road_info', {})
        lanes = road_info.get('lanes', 2)
        length_km = float(road_info.get('length_km', 1.0))
        total_capacity = road_info.get('total_capacity', 3000)
        free_flow_time = road_info.get('free_flow_time_minutes', 10)
        disruption_factors = road_info.get('disruption_factors', {})
        
        # Get disruption impact factor
        impact_factor = disruption_factors.get(disruption_type, 0.6)
        
        # Generate hourly predictions
        hourly_predictions = []
        current_datetime = start_datetime
        
        while current_datetime <= end_datetime:
            # Prepare input
            hour_input = {
                'date': current_datetime.strftime('%Y-%m-%d'),
                'hour': current_datetime.hour,
                'area': area,
                'road_corridor': road_corridor,
                'has_disruption': 1,
                'disruption_type': disruption_type,
                'total_volume': data.get('total_volume', 0),
                'has_real_status': 0
            }
            
            # Make prediction
            prediction = predictor.predict(hour_input)
            
            # IMPORTANT: Calculate delay with road-specific info
            delay_info = predictor.estimate_delay(
                severity=prediction['severity'],
                base_travel_time_minutes=free_flow_time,
                road_length_km=length_km,
                impact_factor=impact_factor
            )
            
            # Add to results
            hourly_predictions.append({
                'datetime': current_datetime.strftime('%Y-%m-%d %H:%M'),
                'date': current_datetime.strftime('%Y-%m-%d'),
                'hour': current_datetime.hour,
                'day_of_week': current_datetime.strftime('%A'),
                'severity': prediction['severity'],
                'severity_label': prediction['severity_label'],
                'confidence': round(prediction['confidence'], 3),
                'delay_info': delay_info,  # ← THIS IS THE KEY LINE
                'probabilities': {
                    k: round(v, 3) for k, v in prediction['probabilities'].items()
                }
            })
            
            current_datetime += timedelta(hours=1)
        
        # Calculate summary statistics
        total_hours = len(hourly_predictions)
        light_hours = sum(1 for p in hourly_predictions if p['severity'] == 0)
        moderate_hours = sum(1 for p in hourly_predictions if p['severity'] == 1)
        heavy_hours = sum(1 for p in hourly_predictions if p['severity'] == 2)
        avg_severity = sum(p['severity'] for p in hourly_predictions) / total_hours
        avg_delay = sum(p['delay_info']['additional_delay_min'] for p in hourly_predictions) / total_hours
        
        # Time segment breakdown
        time_segments = {
            'morning': {'light': 0, 'moderate': 0, 'heavy': 0},
            'afternoon': {'light': 0, 'moderate': 0, 'heavy': 0},
            'night': {'light': 0, 'moderate': 0, 'heavy': 0}
        }
        
        for pred in hourly_predictions:
            hour = pred['hour']
            severity_label = pred['severity_label'].lower()
            
            if 6 <= hour <= 11:
                time_segments['morning'][severity_label] += 1
            elif 12 <= hour <= 17:
                time_segments['afternoon'][severity_label] += 1
            else:
                time_segments['night'][severity_label] += 1
        
        simulation_id = f"sim_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        return jsonify({
            'success': True,
            'simulation_id': simulation_id,
            'input': {
                'area': area,
                'road_corridor': road_corridor,
                'disruption_type': disruption_type,
                'start': start_datetime.strftime('%Y-%m-%d %H:%M'),
                'end': end_datetime.strftime('%Y-%m-%d %H:%M'),
                'description': data.get('description', ''),
                'coordinates': data.get('coordinates', {}),
                'road_info': road_info
            },
            'summary': {
                'total_hours': total_hours,
                'duration_days': round((end_datetime - start_datetime).total_seconds() / 86400, 1),
                'light_hours': light_hours,
                'moderate_hours': moderate_hours,
                'heavy_hours': heavy_hours,
                'light_percentage': round(light_hours / total_hours * 100, 1),
                'moderate_percentage': round(moderate_hours / total_hours * 100, 1),
                'heavy_percentage': round(heavy_hours / total_hours * 100, 1),
                'avg_severity': round(avg_severity, 2),
                'avg_severity_label': 'Light' if avg_severity < 0.5 else ('Moderate' if avg_severity < 1.5 else 'Heavy'),
                'avg_delay_minutes': round(avg_delay, 1),
                'total_delay_hours': round(sum(p['delay_info']['additional_delay_min'] for p in hourly_predictions) / 60, 1)
            },
            'hourly_predictions': hourly_predictions,
            'time_segments': time_segments
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

# ============================================================
# ROUTE 7: Get Recommendations (Keep this)
# ============================================================

@app.route('/api/get-recommendations', methods=['POST'])
def get_recommendations():
    """Get mitigation recommendations"""
    data = request.get_json()
    disruption_type = data.get('disruption_type', 'roadwork')
    avg_severity = data.get('avg_severity', 1.0)
    heavy_percentage = data.get('heavy_percentage', 0)
    
    recommendations = []
    
    if avg_severity > 1.5:
        recommendations.append({
            'priority': 'high',
            'category': 'scheduling',
            'recommendation': 'Consider rescheduling to off-peak hours or weekends',
            'reason': 'High average congestion severity predicted'
        })
    
    if heavy_percentage > 30:
        recommendations.append({
            'priority': 'high',
            'category': 'traffic_management',
            'recommendation': 'Deploy traffic enforcers during peak hours',
            'reason': f'{heavy_percentage}% of hours will have heavy congestion'
        })
    
    if disruption_type == 'roadwork':
        recommendations.append({
            'priority': 'medium',
            'category': 'communication',
            'recommendation': 'Post advance notices 1 week before start date',
            'reason': 'Allow commuters to plan alternate routes'
        })
        
    elif disruption_type == 'event':
        recommendations.append({
            'priority': 'high',
            'category': 'traffic_management',
            'recommendation': 'Implement temporary one-way traffic scheme',
            'reason': 'Manage event-related traffic flow'
        })
    
    return jsonify({
        'success': True,
        'recommendations': recommendations
    })

# ============================================================
# (Data/Uploaded) ROUTE 1: List All Files
# ============================================================

@app.route('/api/files/list', methods=['GET'])
def list_files():
    """
    Get list of all CSV files in data/final folder
    Returns file metadata including name, size, rows, cols, last modified
    """
    try:
        files = []
        
        if not os.path.exists(UPLOAD_FOLDER):
            return jsonify({
                'success': True,
                'files': []
            })
        
        for filename in os.listdir(UPLOAD_FOLDER):
            if filename.endswith('.csv'):
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                
                # Get file stats
                file_stats = os.stat(filepath)
                file_size = file_stats.st_size
                modified_time = datetime.fromtimestamp(file_stats.st_mtime)
                
                # Get CSV info (rows/cols)
                csv_info = get_csv_info(filepath)
                
                files.append({
                    'id': filename,  # Using filename as ID
                    'name': filename,
                    'size_bytes': file_size,
                    'size_mb': round(file_size / (1024 * 1024), 2),
                    'rows': csv_info['rows'],
                    'cols': csv_info['cols'],
                    'updated': modified_time.strftime('%Y-%m-%d'),
                    'updated_full': modified_time.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        # Sort by most recent first
        files.sort(key=lambda x: x['updated_full'], reverse=True)
        
        return jsonify({
            'success': True,
            'files': files,
            'count': len(files)
        })
        
    except Exception as e:
        print(f"Error listing files: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# (Data/Uploaded) ROUTE 2: Upload File
# ============================================================

@app.route('/api/files/upload', methods=['POST'])
def upload_file():
    """
    Upload a CSV file to data/final folder
    """
    try:
        # Check if file is in request
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided'
            }), 400
        
        file = request.files['file']
        
        # Check if file was selected
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400
        
        # Validate file type
        if not allowed_csv_file(file.filename):
            return jsonify({
                'success': False,
                'error': 'Only CSV files are allowed'
            }), 400
        
        # Secure the filename
        filename = secure_filename(file.filename)
        
        # Check if file already exists
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            return jsonify({
                'success': False,
                'error': f'File "{filename}" already exists. Please rename or delete the existing file first.'
            }), 409
        
        # Save file
        file.save(filepath)
        
        # Get file info
        csv_info = get_csv_info(filepath)
        file_stats = os.stat(filepath)
        
        return jsonify({
            'success': True,
            'message': f'File "{filename}" uploaded successfully',
            'file': {
                'name': filename,
                'rows': csv_info['rows'],
                'cols': csv_info['cols'],
                'size_mb': round(file_stats.st_size / (1024 * 1024), 2),
                'updated': datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        print(f"Error uploading file: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# (Data/Uploaded) ROUTE 3: Download Single File
# ============================================================

@app.route('/api/files/download/<filename>', methods=['GET'])
def download_file(filename):
    """
    Download a specific file
    """
    try:
        # Secure the filename
        filename = secure_filename(filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Check if file exists
        if not os.path.exists(filepath):
            return jsonify({
                'success': False,
                'error': 'File not found'
            }), 404
        
        # Send file
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# (Data/Uploaded) ROUTE 4: Download Multiple Files (as ZIP)
# ============================================================

@app.route('/api/files/download-multiple', methods=['POST'])
def download_multiple_files():
    """
    Download multiple files as a ZIP archive
    """
    try:
        import zipfile
        import io
        
        data = request.get_json()
        filenames = data.get('filenames', [])
        
        if not filenames:
            return jsonify({
                'success': False,
                'error': 'No files selected'
            }), 400
        
        # Create in-memory ZIP file
        memory_file = io.BytesIO()
        
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename in filenames:
                filename = secure_filename(filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                if os.path.exists(filepath):
                    zf.write(filepath, arcname=filename)
        
        memory_file.seek(0)
        
        # Generate ZIP filename with timestamp
        zip_filename = f'urbanflow_datasets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        
        return send_file(
            memory_file,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        print(f"Error creating ZIP: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# (Data/Uploaded) ROUTE 5: Delete File(s)
# ============================================================

@app.route('/api/files/delete', methods=['POST'])
def delete_files():
    """
    Delete one or more files
    """
    try:
        data = request.get_json()
        filenames = data.get('filenames', [])
        
        if not filenames:
            return jsonify({
                'success': False,
                'error': 'No files selected for deletion'
            }), 400
        
        deleted = []
        failed = []
        
        for filename in filenames:
            try:
                filename = secure_filename(filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted.append(filename)
                else:
                    failed.append({
                        'filename': filename,
                        'reason': 'File not found'
                    })
            except Exception as e:
                failed.append({
                    'filename': filename,
                    'reason': str(e)
                })
        
        return jsonify({
            'success': len(failed) == 0,
            'deleted': deleted,
            'failed': failed,
            'message': f'Successfully deleted {len(deleted)} file(s)'
        })
        
    except Exception as e:
        print(f"Error deleting files: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# (Data/Uploaded) ROUTE 6: Get File Preview (first 10 rows)
# ============================================================

@app.route('/api/files/preview/<filename>', methods=['GET'])
def preview_file(filename):
    """
    Get preview of CSV file (first 10 rows)
    """
    try:
        filename = secure_filename(filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath):
            return jsonify({
                'success': False,
                'error': 'File not found'
            }), 404
        
        # Read CSV
        df = pd.read_csv(filepath)
        
        # Get first 10 rows
        preview_data = df.head(10).to_dict('records')
        columns = df.columns.tolist()
        
        return jsonify({
            'success': True,
            'filename': filename,
            'columns': columns,
            'preview': preview_data,
            'total_rows': len(df)
        })
        
    except Exception as e:
        print(f"Error previewing file: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# OTP GENERATION AND VERIFICATION
# ============================================================

@app.route('/api/send-publish-otp', methods=['POST'])
def send_publish_otp():
    """
    Generate and send OTP for publishing verification
    In production, this would send an email. For now, we'll return it in response for testing.
    """
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        simulation_id = data.get('simulation_id')
        
        if not simulation_id:
            return jsonify({
                'success': False,
                'error': 'simulation_id is required'
            }), 400
        
        # Generate 6-digit OTP
        otp_code = ''.join(random.choices(string.digits, k=6))
        
        # Initialize variables to avoid linting warnings
        user_email = None
        user_name = None
        # Store OTP in database
        conn = None
        try:
            conn = db._get_connection()
            cursor = conn.cursor()
            
            # Set expiry to 10 minutes from now
            expires_at = datetime.now() + timedelta(minutes=10)
            
            cursor.execute("""
                INSERT INTO verification_otps (user_id, simulation_id, otp_code, expires_at)
                VALUES (%s, %s, %s, %s)
            """, (user_id, simulation_id, otp_code, expires_at))
            
            # Get user email
            cursor.execute("""
                SELECT email, 
                       CONCAT(first_name, ' ', last_name) as full_name 
                FROM users 
                WHERE user_id = %s
            """, (user_id,))
            user = cursor.fetchone()
            
            if not user:
                return jsonify({
                    'success': False,
                    'error': 'User not found'
                }), 404
            
            user_email, user_name = user

            conn.commit()
            
            # Send actual email
            try:
                send_otp_email(user_email, user_name, otp_code, expires_minutes=10)
                print(f"✓ OTP email sent to {user_email}")
                
                return jsonify({
                    'success': True,
                    'message': f'Verification code sent to {user_email}',
                    'expires_in_minutes': 10
                })
                
            except Exception as email_error:
                # Rollback OTP if email fails
                cursor.execute(
                    "DELETE FROM verification_otps WHERE simulation_id = %s AND otp_code = %s",
                    (simulation_id, otp_code)
                )
                conn.commit()
                
                return jsonify({
                    'success': False,
                    'error': 'Failed to send verification email. Please try again.',
                    'details': str(email_error)
                }), 500
            
        finally:
            if conn:
                conn.close()
                
    except Exception as e:
        print(f"✗ Error sending OTP: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'An error occurred while processing your request'
        }), 500

@app.route('/api/verify-publish-otp', methods=['POST'])
def verify_publish_otp():
    """
    Verify OTP and publish simulation if valid
    """
    try:
        data = request.get_json()
        
        simulation_id = data.get('simulation_id')
        otp_code = data.get('otp_code')
        user_id = data.get('user_id')
        title = data.get('title')
        public_description = data.get('public_description')
        
        if not simulation_id or not otp_code:
            return jsonify({
                'success': False,
                'error': 'simulation_id and otp_code are required'
            }), 400
        
        conn = None
        try:
            conn = db._get_connection()
            cursor = conn.cursor()
            
            # Verify OTP
            cursor.execute("""
                SELECT otp_id, expires_at, is_used
                FROM verification_otps
                WHERE simulation_id = %s 
                  AND otp_code = %s 
                  AND user_id = %s
                ORDER BY created_at DESC
                LIMIT 1
            """, (simulation_id, otp_code, user_id))
            
            otp_record = cursor.fetchone()
            
            if not otp_record:
                return jsonify({
                    'success': False,
                    'error': 'Invalid OTP code'
                }), 400
            
            otp_id, expires_at, is_used = otp_record
            
            # Check if already used
            if is_used:
                return jsonify({
                    'success': False,
                    'error': 'OTP code has already been used'
                }), 400
            
            # Check if expired
            if datetime.now() > expires_at:
                return jsonify({
                    'success': False,
                    'error': 'OTP code has expired'
                }), 400
            
            # Mark OTP as used
            cursor.execute("""
                UPDATE verification_otps
                SET is_used = TRUE, used_at = CURRENT_TIMESTAMP
                WHERE otp_id = %s
            """, (otp_id,))
            
            # Publish the simulation
            slug = db.publish_simulation(
                simulation_id=simulation_id,
                published_by_user_id=user_id,
                title=title,
                public_description=public_description
            )
            
            if slug:
                return jsonify({
                    'success': True,
                    'slug': slug,
                    'message': 'Simulation published successfully',
                    'public_url': f'/disruptions/{slug}'
                })
            else:
                return jsonify({
                    'success': False,
                    'error': 'Failed to publish simulation'
                }), 500
            
        finally:
            if conn:
                conn.close()
                
    except Exception as e:
        print(f"✗ Error verifying OTP: {e}")
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================
# BATCH DELETE SIMULATIONS
# ============================================================

@app.route('/api/delete-simulations-batch', methods=['POST'])
def delete_simulations_batch():
    """
    Delete multiple simulations at once
    """
    try:
        data = request.get_json()
        simulation_ids = data.get('simulation_ids', [])
        user_id = data.get('user_id')
        
        if not simulation_ids:
            return jsonify({
                'success': False,
                'error': 'No simulation IDs provided'
            }), 400
        
        deleted_count = 0
        failed_ids = []
        
        for sim_id in simulation_ids:
            success = db.delete_simulation(sim_id, user_id)
            if success:
                deleted_count += 1
            else:
                failed_ids.append(sim_id)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'failed_count': len(failed_ids),
            'failed_ids': failed_ids,
            'message': f'Successfully deleted {deleted_count} simulation(s)'
        })
        
    except Exception as e:
        print(f"✗ Error deleting simulations: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500





@app.route('/api/validate-calculations', methods=['POST'])
def validate_calculations():
    """
    Validate traffic calculations against known benchmarks
    Returns comparison with expected values from literature
    """
    data = request.get_json()
    
    results = {
        'capacity_validation': {},
        'delay_validation': {},
        'disruption_validation': {}
    }
    
    # Test capacity calculation
    test_cases = [
        {'type': 'motorway', 'speed': 100, 'expected_range': (2200, 2400)},
        {'type': 'primary', 'speed': 60, 'expected_range': (1600, 1800)},
        {'type': 'residential', 'speed': 30, 'expected_range': (900, 1100)}
    ]
    
    for case in test_cases:
        calculated = calculate_lane_capacity(case['type'], case['speed'])
        expected_min, expected_max = case['expected_range']
        within_range = expected_min <= calculated <= expected_max
        
        results['capacity_validation'][case['type']] = {
            'calculated': calculated,
            'expected_range': case['expected_range'],
            'valid': within_range
        }
    
    return jsonify({
        'success': True,
        'results': results,
        'references': {
            'capacity': 'Highway Capacity Manual 2016',
            'delay': 'BPR Function (1964)',
            'disruption': 'FHWA Work Zone Studies (2011)'
        }
    })

# ============================================================
# RUN APP
# ============================================================

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)